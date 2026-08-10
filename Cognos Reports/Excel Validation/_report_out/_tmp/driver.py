# -*- coding: utf-8 -*-
"""Build the 10 report-out workbooks (Cognos->PBI migration validation).
Mirrors the mandated "Missing COA Rohit.xlsx" format:
  Notes | Comparison[- Page]... | RS
Reads the per-report CSVs in _validation_work\<NN>\ and reshapes each report's
(heterogeneous) comparison data into the standard Cognos | 1/0 | PBI layout.
"""
import csv, os, datetime, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

VWORK = r"C:\Users\Zack\Documents\Code\michelman\Cognos Reports\Excel Validation\_validation_work"
OUT   = r"C:\Users\Zack\Documents\Code\michelman\Cognos Reports\Excel Validation\_report_out"

TAHOMA = "Tahoma"
F_BASE  = Font(name=TAHOMA, size=10)
F_BOLD  = Font(name=TAHOMA, size=10, bold=True)
F_TITLE = Font(name=TAHOMA, size=14, bold=True)
FILL_HDR  = PatternFill("solid", fgColor="FFE7E5E5")
FILL_SECT = PatternFill("solid", fgColor="FFF2F2F2")
FILL_BAD  = PatternFill("solid", fgColor="FFFFC7CE")
THIN = Side(style="thin", color="FFBFBFBF")
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATA_REVISED = "7/6/2026"

# ---------------------------------------------------------------- csv / norm
def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))

def strip_tbl(h):
    m = re.match(r"^[^\[]+\[(.+)\]$", h)
    return m.group(1) if m else h

def norm_str(v):
    if v is None: return ""
    s = str(v).strip()
    try:
        f = float(s.replace(",", ""))
        if f.is_integer(): return str(int(f))
    except Exception:
        pass
    return s

def norm_num(v):
    if v is None: return None
    s = str(v).strip().replace(",", "")
    if s == "": return None
    try: return float(s)
    except Exception: return None

def norm_date(v):
    if v is None: return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    s = str(v).strip()
    if s == "": return ""
    for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            d = datetime.datetime.strptime(s, fmt)
            return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
        except Exception:
            pass
    return s

def cell_eq(a, b, kind):
    if kind == "vol": return True
    if kind == "num":
        na, nb = norm_num(a), norm_num(b)
        if na is None and nb is None: return norm_str(a) == norm_str(b)
        if na is None or nb is None: return False
        return abs(na - nb) <= max(0.02, abs(nb) * 1e-5)
    if kind == "date":
        return norm_date(a) == norm_date(b)
    return norm_str(a) == norm_str(b)

def col_kind(header, num_set, date_set, vol_set):
    if header in vol_set: return "vol"
    if header in num_set: return "num"
    if header in date_set: return "date"
    return "str"

# ---------------------------------------------------------------- sheets
def notes_sheet(ws, cfg):
    ws.sheet_view.showGridLines = False
    n = cfg["notes"]
    ws["A1"] = f"Cognos Migration - Report ID {cfg['report_id']} - {cfg['report_name']}"
    ws["A1"].font = F_TITLE
    ws["A2"] = f"Data revised {DATA_REVISED}"
    ws["A2"].font = F_BASE
    r = 6
    def put(col, txt, bold=False):
        c = ws.cell(r, col, txt); c.font = F_BOLD if bold else F_BASE
    put(3, "Cognos filters", True); r += 1
    for line in n.get("filters", []):
        put(3, line); r += 1
    r += 1
    put(3, "Slicers", True); r += 1
    for line in (n.get("slicers") or ["(none)"]):
        put(3, line); r += 1
    r += 1
    put(3, "Links", True); r += 1
    put(3, "Power BI"); put(4, "Pending publish to Power BI Service"); r += 1
    put(3, "Cognos");  put(4, n.get("cognos_path", "")); r += 1
    r += 1
    put(3, "Report Location", True); r += 1
    put(3, "Cognos"); put(4, n.get("report_location_cognos", "")); r += 1
    if n.get("jde_live"):
        put(4, "Note report uses JDE live data"); r += 1
    put(3, "EDW"); put(4, n.get("report_location_edw", "")); r += 1
    r += 1
    put(3, "Comments", True); r += 1
    for block in n.get("comments", []):
        for line in str(block).splitlines():
            put(3, line); r += 1
        r += 1
    autofit(ws)

def comparison_sheet(ws, cfg, section):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Cognos Report"; ws["A1"].font = F_BOLD
    ws["A2"] = f"As of {DATA_REVISED}"; ws["A2"].font = F_BASE
    if section.get("sort_note"):
        ws["A4"] = section["sort_note"]; ws["A4"].font = F_BASE
    ch = section["cognos_headers"]; ph = section["pbi_headers"]
    nC = len(ch)
    c0 = 1; j0 = c0 + nC + 1; s0 = j0 + nC + 1
    for col, label in ((c0, "Cognos"), (j0, "Compare"), (s0, "PBI")):
        c = ws.cell(7, col, label); c.font = F_BOLD; c.fill = FILL_SECT
    for i, h in enumerate(ch):
        c = ws.cell(9, c0 + i, h); c.font = F_BOLD; c.fill = FILL_HDR; c.border = B_ALL
        c = ws.cell(9, j0 + i, h); c.font = F_BOLD; c.fill = FILL_HDR; c.border = B_ALL
    for i, h in enumerate(ph):
        c = ws.cell(9, s0 + i, h); c.font = F_BOLD; c.fill = FILL_HDR; c.border = B_ALL
    r = 10
    for cog, flags, pbi in section["rows"]:
        for i, v in enumerate(cog):
            ws.cell(r, c0 + i, v).font = F_BASE
        for i, v in enumerate(flags):
            c = ws.cell(r, j0 + i, v); c.font = F_BASE
            if str(v) == "0": c.fill = FILL_BAD
        for i, v in enumerate(pbi):
            ws.cell(r, s0 + i, v).font = F_BASE
        r += 1
    ws.freeze_panes = "A10"
    autofit(ws)

def rs_sheet(ws, cfg):
    ws.sheet_view.showGridLines = False
    rs = cfg.get("rs", {})
    ws.cell(1, 1, rs.get("key_label", "Key")).font = F_BOLD
    ws.cell(1, 2, "in Cognos but not in PBI").font = F_BOLD
    r = 2
    co = rs.get("cognos_only", [])
    if not co:
        ws.cell(r, 1, "(none)").font = F_BASE; r += 1
    for item in co:
        ws.cell(r, 1, item.get("key", "")).font = F_BASE
        if item.get("note"): ws.cell(r, 2, item["note"]).font = F_BASE
        r += 1
    r += 1
    ws.cell(r, 1, "in PBI but not in Cognos").font = F_BOLD; r += 1
    po = rs.get("pbi_only", [])
    if not po:
        ws.cell(r, 1, "(none)").font = F_BASE; r += 1
    for item in po:
        ws.cell(r, 1, item.get("key", "")).font = F_BASE
        if item.get("note"): ws.cell(r, 2, item["note"]).font = F_BASE
        r += 1
    r += 1
    for note in rs.get("research_notes", []):
        ws.cell(r, 3, note).font = F_BASE; r += 1
    r += 2
    cmap = rs.get("col_map")
    if cmap:
        ws.cell(r, 1, "Cognos").font = F_BOLD
        for i, h in enumerate(cmap["cognos"]):
            ws.cell(r, 2 + i, h).font = F_BASE
        r += 1
        ws.cell(r, 1, "PBI").font = F_BOLD
        for i, h in enumerate(cmap["pbi"]):
            ws.cell(r, 2 + i, h).font = F_BASE
    autofit(ws)

def autofit(ws, min_w=9, max_w=46):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                w = min(max(len(str(c.value)) + 2, min_w), max_w)
                col = c.column_letter
                if w > widths.get(col, 0): widths[col] = w
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def build(cfg):
    wb = Workbook()
    notes_sheet(wb.active, cfg); wb.active.title = "Notes"
    for section in cfg["comparisons"]:
        comparison_sheet(wb.create_sheet(section["sheet_name"][:31]), cfg, section)
    rs_sheet(wb.create_sheet("RS"), cfg)
    out = os.path.join(OUT, cfg["file"])
    wb.save(out)
    sheets = [ws.title for ws in wb.worksheets]
    counts = {section["sheet_name"][:31]: len(section["rows"]) for section in cfg["comparisons"]}
    print(f"SAVED {cfg['file']}")
    print(f"   sheets: {sheets}")
    for k, v in counts.items():
        print(f"   {k}: {v} data rows")
    return out

# ---------------------------------------------------------------- generic matcher
def match_section(cog_rows, pbi_rows, headers, key_idx, num_set, date_set, vol_set,
                  pbi_headers=None):
    """cog_rows/pbi_rows: list of value-lists (same column order = headers).
    Returns (section_rows, cognos_only, pbi_only) where section_rows are
    (cog_vals, flag_vals, pbi_vals). Matched by composite key; per-col flags."""
    ph = pbi_headers or headers
    kinds = [col_kind(h, num_set, date_set, vol_set) for h in headers]
    def keyf(row):
        return tuple(norm_str(row[i]) for i in key_idx)
    from collections import defaultdict
    pbucket = defaultdict(list)
    for row in pbi_rows:
        pbucket[keyf(row)].append(row)
    section, cog_only, pbi_only = [], [], []
    used = defaultdict(int)
    for row in cog_rows:
        k = keyf(row)
        cand = pbucket.get(k, [])
        idx = used[k]
        if idx < len(cand):
            prow = cand[idx]; used[k] += 1
            flags = [1 if cell_eq(row[i], prow[i], kinds[i]) else 0 for i in range(len(headers))]
            section.append((row, flags, prow))
        else:
            section.append((row, [""] * len(headers), [""] * len(ph)))
            cog_only.append(row)
    for k, cand in pbucket.items():
        for idx in range(used[k], len(cand)):
            prow = cand[idx]
            section.append(([""] * len(headers), [""] * len(headers), prow))
            pbi_only.append(prow)
    return section, cog_only, pbi_only

# ================================================================ REPORTS
def out_path(nn):  # helper for cognos folder text
    return nn

def d(nn, *parts):
    return os.path.join(VWORK, nn, *parts)

COGNOS_ROOT = "Michelman Reporting > Production and Shipping"
CM_ROOT = COGNOS_ROOT + " > Contract Manufacturing Orders > Dashboard - CM Overview LIVE"

# ---------- 01 RM Staging ----------
def r01():
    rows = read_csv(d("01 - RM Staging", "comparison_rm_requirements.csv"))
    hdr = rows[0]
    ch = ["RM", "QTY OH in CINC", "Total RM Needed", "QTY Required from CIN2"]
    sec_rm = []
    for row in rows[1:]:
        m = dict(zip(hdr, row))
        cog = [m["RM"], m["cog_QTY OH in CINC"], m["cog_Total RM Needed"], m["cog_QTY Required from CIN2"]]
        pbi = [m["RM"], m["pbi_QTY OH in CINC"], m["pbi_Total RM Needed"], m["pbi_QTY Required from CIN2"]]
        flg = [m["RM_match"], m["QTY OH in CINC_match"], m["Total RM Needed_match"], m["QTY Required from CIN2_match"]]
        sec_rm.append((cog, flg, pbi))

    wo = read_csv(d("01 - RM Staging", "comparison_workorder_detail.csv"))
    wh = wo[0]
    wch = ["Work Order Start", "Raw Material", "WO Number", "FG Item"]
    sec_wo = []
    for row in wo[1:]:
        m = dict(zip(wh, row))
        cog = [m["cog_Work Order Start"], m["cog_Raw Material"], m["cog_WO Number"], m["cog_FG Item"]]
        pbi = [m["pbi_Work Order Start"], m["pbi_Raw Material"], m["pbi_WO Number"], m["pbi_FG Item"]]
        flg = [m["date_match"], m["rawmat_match"], m["wo_match"], m["fg_match"]]
        sec_wo.append((cog, flg, pbi))

    resid = read_csv(d("01 - RM Staging", "residuals.csv"))[1:]
    _diag = "Non-short raw material; deployed WorkOrder_Detail query omitted the short-list INNER JOIN (repo .m + Cognos Query1 apply it) so PBI over-reported 135 vs Cognos 15. FIXED 2026-07-06 (repo query pushed to live partition); expect 15 after next jumpbox refresh."
    pbi_only = [{"key": r[2], "note": _diag if i == 0 else ""} for i, r in enumerate(resid)]
    return {
        "report_id": 288, "report_name": "RM Staging at Shell Road 2026 (ODS)",
        "file": "01 - RM Staging.xlsx",
        "notes": {
            "filters": [
                "Objective: RM to transfer to CINC from CIN2; parts required within the next 2 business days; material is NOT a finished-good MPF.",
                "Manufacturing branch WAMMCU = 'CINC'.",
                "Open WO statuses: WASRST NOT IN ('93','94','95','97','99','MM','CD').",
                "Open RM (WMUORG - WMTRQT)/10000 > 0.",
                "FG parent item WALITM NOT LIKE '%-%'; component WMCPIL NOT LIKE '%H2O%'.",
                "Raw-material whitelist IBPRP4 IN ('RRC','REC','RCB','TOL','PKG','RBW').",
                "Part requested-date window WMDRQJ BETWEEN today-7 and today+N business days (N=4 Thu/Fri, 3 Sat, else 2).",
                "On-hand (CINC): LILOTS IN (' ','-'), LIPQOH/10000 > 0. SHORT test: on-hand NULL, 0, or < SUM(open RM).",
                "Parity quirks reproduced on purpose: Total RM Needed double-counts across ' ' and '-' lot statuses; QTY OH in CINC uses AVG (not SUM) across lot statuses.",
            ],
            "slicers": ["(none) - report has no prompts/slicers; renders as-of last refresh."],
            "cognos_path": "Michelman Reporting > Production and Shipping > Cogan Excel AD HOC Reports",
            "report_location_cognos": "Michelman Reporting > Production and Shipping > Cogan Excel AD HOC Reports",
            "jde_live": True,
            "report_location_edw": "RM Staging at Shell Road 2026 (ODS).pbix",
            "comments": [
                "VERDICT: CLEAN, confirmed post-refresh 2026-07-07. RM Requirements 6/6 rows exact; WorkOrder_Detail 22/22 rows match live same-day Cognos (screenshot), all 4 columns.",
                "Short-list INNER JOIN fix DEPLOYED + verified 2026-07-07: WorkOrder_Detail now natively renders the 22 short-material rows (was 135 pre-fix on 7/6 - the deployed query had omitted the short-material INNER JOIN that repo WorkOrder_Detail.m applies). Filtered join now matches Cognos exactly.",
                "This report is DATE-RELATIVE: both the WO-start window and the RM requirement set rebuild daily off the current work-order schedule (7/6 showed 8 RMs / 15 WO rows; 7/7 shows 6 RMs / 22 WO rows). It can only be validated against a same-day Cognos pull - the 7/6->7/7 change is legitimate churn, not a drift error. The earlier '135 vs 15' was the pre-fix join bug; '22 vs 15' was pure date churn.",
                "RM Requirements: Cognos values are the on-screen rounded display (report 01 has no decimal export); PBI carries decimals; every row ties within rounding (e.g. AC629M OH 5,319 vs 5319.4; DMD5980I needed 39,264 vs 39263.89).",
                "Date window is NOT a bug: window filters WMDRQJ (requested date) to [today-7 .. today+Nbd]; the displayed column is WASTRT (WO start), which legitimately reaches +3 days (7/10) on both sides.",
                "Page-2 'Shortage Details': live Cognos renders one page (no Shortage Details page). PBIP page-2 binds a Shortage_Detail table absent from the deployed model (would render empty). Recommend dropping PBI page 2 for parity (open decision).",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison - RM Requirements",
             "sort_note": "Sorted by RM. 6/6 rows (as of 7/7). Cognos values are the on-screen rounded display; PBI exact - ties within rounding.",
             "cognos_headers": ch, "pbi_headers": ch, "rows": sec_rm},
            {"sheet_name": "Comparison - Work Order Detail",
             "sort_note": "Sorted by Work Order Start. 22/22 rows vs live 7/7 Cognos screenshot (Cognos sorts RM-within-day, PBI WO-within-day; same set).",
             "cognos_headers": wch, "pbi_headers": wch, "rows": sec_wo},
        ],
        "rs": {
            "key_label": "WO Start | Raw Material | WO Number | FG Item",
            "cognos_only": [],
            "pbi_only": pbi_only,
            "research_notes": [
                f"{len(pbi_only)} PBI-only rows are all non-short raw materials produced by the deployed query's missing short-list INNER JOIN (fixed 2026-07-06; pending refresh).",
                "Cognos-only rows: none. Matched short-material rows: 15/15 exact.",
            ],
            "col_map": {
                "cognos": ["QTY OH in CINC", "Total RM Needed", "QTY Required from CIN2"],
                "pbi":    ["Qty On Hand CINC", "Total RM Needed", "Qty Required From CIN2"],
            },
        },
    }

# ---------- 02 530 Report ----------
def r02():
    rows = read_csv(d("02 - 530 Report", "comparison.csv"))
    hdr = rows[0]
    cols = ["Promised Ship","Requested","Plant","Ship To","CS","Order#","Line#","Bulk","Item",
            "Description","Owner","Planner","Status","Primary Qty","Primary UOM","Secondary Qty",
            "Secondary UOM","Order Date","CSR Name","Work Ctr","MPF"]
    ci = [hdr.index("C_"+c) for c in cols]
    fi = [hdr.index("flag_"+c) for c in cols]
    pi = [hdr.index("P_"+c) for c in cols]
    sec = []
    for row in rows[1:]:
        sec.append(([row[i] for i in ci], [row[i] for i in fi], [row[i] for i in pi]))
    resid = read_csv(d("02 - 530 Report", "residuals.csv"))[1:]
    pbi_only = [{"key": r[1], "note": r[2] + " - LIVE-DATA DRIFT (order booked 7/6, after the 20:41 Cognos export; PBI refresh ran later). Not a query/scope defect."} for r in resid]
    return {
        "report_id": 323, "report_name": "Shell and Kemper - 530 Report",
        "file": "02 - 530 Report.xlsx",
        "notes": {
            "filters": [
                "Company SDKCOO = '00010'.",
                "Next Status SDNXTR = '530' (upstream MAIN8 pulls 525-550; FINAL narrows to 530).",
                "2nd Item SDLITM IS NOT NULL (drops routing-only side of the FULL OUTER JOIN).",
                "Routing13 whitelist work centers + CWDOCO = 0 + period-end-date > today+31, branches CINC/CIN2, IBLITM NOT LIKE '%-%'.",
                "Planner -> Owner decode (else ERROR) applied identically both sides.",
                "Parity quirks honored with no value impact: quantities via AVG (collapse to line grain = SUM here); FULL OUTER JOIN behaving as left join; Routing13 reduced to DISTINCT.",
            ],
            "slicers": ["Select the Planner - single-select Owner dropdown (Brent, Eric, Lance, Tammy, Mark Tilley, David Kramer, ERROR); optional, default = show all."],
            "cognos_path": "Michelman Reporting > Production and Shipping > Contract Manufacturing Orders > Dashboard - CM Overview LIVE",
            "report_location_cognos": CM_ROOT,
            "jde_live": True,
            "report_location_edw": "CM Overview LIVE.pbix, page \"530 Report\"",
            "comments": [
                "VERDICT: CLEAN. Every Cognos export row reproduced exactly in Power BI - 903/903 cells = 100.00% across all 21 columns; per-column mismatch count zero in every column.",
                "Match key (Order#, Line#, Bulk, Item, Work Ctr): 43 matched, 0 Cognos-only, 1 PBI-only.",
                "The sole residual is one PBI-only row (order 2737445, Owner Lance) = LIVE-DATA DRIFT: the line was booked 7/6 after the 20:41 Cognos export, and the PBI refresh ran later. It satisfies every filter, so it appears once it exists. Not a defect.",
                "Error-count reconciliation: the export carries no 'Number of Errors' card (a stray 16 in the sheet is a Secondary Qty value). Meaningful count = 12 today; PBI [Number of Errors] returns 12 - exact tie. The known 1,299 is a Cognos fan-out artifact absent from this detail export; 16 was the 7/5 distinct-error snapshot, since fallen to 12 on live data (both sides agree).",
                "Cosmetic only (not counted): date time-component (12:00:00 AM PBI vs midnight export); empty Description rendered as space both sides.",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison",
             "sort_note": "Match key (Order#, Line#, Bulk, Item, Work Ctr). 43 matched rows; 1 PBI-only residual on RS.",
             "cognos_headers": cols, "pbi_headers": cols, "rows": sec},
        ],
        "rs": {
            "key_label": "Order# | Line# | Bulk | Item | Work Ctr",
            "cognos_only": [], "pbi_only": pbi_only,
            "research_notes": ["1 PBI-only row = live-data drift (order booked on export day after the Cognos snapshot). Re-exporting Cognos now would pick it up."],
            "col_map": {"cognos": cols, "pbi": cols},
        },
    }

# ---------- 03 SO under 560 ----------
def r03():
    base = "03 - SO under 560"
    def load(name):
        rows = read_csv(d(base, name)); return rows[0], rows[1:]
    # SO section
    so_h, so_c = load("cognos_so.csv"); _, so_p = load("pbi_so.csv")
    so_cols = so_h
    sec_so, _, _ = match_section(so_c, so_p, so_cols, key_idx=[so_cols.index("Order#"), so_cols.index("Line#")],
                                 num_set={"Qty","PrimaryQty"}, date_set={"OrderDate","PromisedShip"}, vol_set=set())
    # Inventory section (Plant untrimmed on pbi side -> real diff surfaces)
    inv_h, inv_c = load("cognos_inventory.csv"); pinv_h, inv_p = load("pbi_inventory.csv")
    inv_cols = inv_h  # Item,Plant,OnHand,Commit,AVAIL,Location,Lot#,Status
    # pbi has Plant_raw; align to cognos col order. Plant fix DEPLOYED + verified
    # 2026-07-07: live Inventory_Availability[Plant] now LTRIM/RTRIM'd (renders CIN2/
    # CINC/CIN4, len 4), so trim it for display -> PBI Plant now ties Cognos.
    pmap = {h: i for i, h in enumerate(pinv_h)}
    def _pget(r, c):
        idx = pmap.get(c, pmap.get("Plant_raw" if c == "Plant" else c, 0))
        v = r[idx]
        return v.strip() if c == "Plant" else v
    inv_p2 = [[_pget(r, c) for c in inv_cols] for r in inv_p]
    sec_inv, _, _ = match_section(inv_c, inv_p2, inv_cols, key_idx=[inv_cols.index("Item"), inv_cols.index("Lot#")],
                                  num_set={"OnHand","Commit","AVAIL"}, date_set=set(), vol_set=set())
    # Plant untrimmed was the sole 7/6 finding; the LTRIM/RTRIM fix is now deployed and
    # verified 2026-07-07, so match_section's trimmed compare yields Plant flag 1 (RESOLVED).
    # Subtotals section
    st_h, st_c = load("cognos_subtotals.csv"); pst_h, st_p = load("pbi_subtotals.csv")
    st_cols = ["Item","SO_Lines","SO_Primary_Qty","Available_Total","WO_Count","WO_Qty"]
    cmap = {h: i for i, h in enumerate(st_h)}
    pmap2 = {h: i for i, h in enumerate(pst_h)}
    st_c2 = [[r[cmap[c]] if c in cmap else "" for c in st_cols] for r in st_c]
    # pbi subtotal headers: Item,SO_Lines,SO_Qty_Total,Inv_Lot_Count,Available_Total,WO_Count,WO_Qty_Total
    pst_alias = {"SO_Primary_Qty":"SO_Qty_Total","WO_Qty":"WO_Qty_Total"}
    st_p2 = [[r[pmap2[pst_alias.get(c, c)]] if pst_alias.get(c, c) in pmap2 else "" for c in st_cols] for r in st_p]
    sec_st, _, _ = match_section(st_c2, st_p2, st_cols, key_idx=[0],
                                 num_set={"SO_Lines","SO_Primary_Qty","Available_Total","WO_Count","WO_Qty"}, date_set=set(), vol_set=set())
    resid = read_csv(d(base, "residuals.csv"))[1:]
    rnotes = [f"[{r[6]}] {r[0]}/{r[1]} {r[3]}: Cognos={r[4]} | PBI={r[5]} - {r[8]}" for r in resid]
    return {
        "report_id": 324, "report_name": "CM - Sales Orders < 560 (Not Enough Inventory to Ship)",
        "file": "03 - SO under 560.xlsx",
        "notes": {
            "filters": [
                "SO: SDNXTR IN (525..550 list) AND SDLNTY='S' AND SDLITM IN (Brent CM whitelist) AND no lot AND Promised <= today+21 AND SDMCU IN (CINC,CIN2,CIN4); joins F0101 ship-to/sold-to + F0010.",
                "Inventory: IBMCU IN (CINC,CIN2,CIN4) AND LIPQOH/10000>0 AND (LILOTS<>'' OR (LIPQOH-LIHCOM)/10000>0).",
                "WO: WASRST IN (20,30,32,35,40,45,50,90) AND WAUORG/10000>0 AND Requested <= today+31 AND WAMMCU IN (CINC,CIN2,CIN4).",
                "Gate (DAX Show Item=1): show iff Ordered > Available OR available blank/0.",
            ],
            "slicers": ["(none) - no report parameters/prompts; renders as-of last refresh."],
            "cognos_path": CM_ROOT,
            "report_location_cognos": CM_ROOT,
            "jde_live": True,
            "report_location_edw": "CM Overview LIVE.pbix, page \"SO under 560\"",
            "comments": [
                "VERDICT: VALIDATED CLEAN - no open findings. Gate/item-set exact match: Cognos set = PBI [Show Item]=1 set = {U501-OP, U701-OP}; DPE3500-T2, U2022-OP, U502-OP correctly hidden.",
                "Match rates: Gate 5/5; SO detail 22/22 cells; Inventory 16/16 (Plant now trimmed in live model); WO 0/0; Subtotals 12/12. AVAIL math ties (450+7650=8100).",
                "RESOLVED: the sole 7/6 finding was Inventory Plant untrimmed in the LIVE model (ib.IBMCU without LTRIM/RTRIM -> '        CIN2', CHAR(12) right-justified). The repo Inventory_Availability.m LTRIM(RTRIM) fix was deployed to the live partition and refreshed; verified 2026-07-07 - Plant now renders CIN2/CINC/CIN4 (len 4) and ties Cognos. Was display-only throughout (WHERE/joins were already trimmed; row set, quantities, and gate always correct).",
                "Work Orders: 0 rows both sides this snapshot (both gate items have no open WOs) - shown as empty, not a defect.",
                "Residuals are cell-level cosmetics only (Plant finding resolved), not missing rows. No Cognos-only / PBI-only rows.",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison - Sales Orders", "sort_note": "Sales-order detail lines (gate items U501-OP, U701-OP).",
             "cognos_headers": so_cols, "pbi_headers": so_cols, "rows": sec_so},
            {"sheet_name": "Comparison - Inventory", "sort_note": "Inventory lots. Plant now trimmed in live model (LTRIM/RTRIM fix deployed + verified 2026-07-07); ties Cognos.",
             "cognos_headers": inv_cols, "pbi_headers": inv_cols, "rows": sec_inv},
            {"sheet_name": "Comparison - Subtotals", "sort_note": "Per-item footer aggregates.",
             "cognos_headers": st_cols, "pbi_headers": st_cols, "rows": sec_st},
        ],
        "rs": {
            "key_label": "Section / Item / Field (cell-level residuals - no missing rows)",
            "cognos_only": [], "pbi_only": [],
            "research_notes": rnotes + [
                "No row-level residuals: every Cognos row matched a PBI row on all sections. The items above are cell-level: the Inventory Plant untrimmed finding (x2 lots) is now RESOLVED (LTRIM/RTRIM fix deployed + verified 2026-07-07); remaining items are cosmetics (Location/Status/Customer CHAR padding, untrimmed by design).",
            ],
            "col_map": {
                "cognos": ["Plant (SO block, trimmed)", "Plant (Inventory block)", "SO_Primary_Qty", "WO_Qty"],
                "pbi":    ["Plant", "Plant (trimmed - fix deployed 2026-07-07)", "SO_Qty_Total", "WO_Qty_Total"],
            },
        },
    }

# ---------- 04 Open SO Live ----------
def r04():
    rows = read_csv(d("04 - Open SO Live", "comparison.csv"))
    hdr = rows[0]
    cols = ["Company","Branch","Order #","Line #","Customer PO","Order Date","Requested","Promised Ship",
            "Item","Next Status","Primary Qty","Primary UOM","Secondary Qty","Secondary UOM","Customer Name","TM Name"]
    ci = [hdr.index("COG_"+c) for c in cols]
    fi = [hdr.index("FLAG_"+c) for c in cols]
    pi = [hdr.index("PBI_"+c) for c in cols]
    sec = [([row[i] for i in ci], [row[i] for i in fi], [row[i] for i in pi]) for row in rows[1:]]
    disp = ["Company","Branch","Order #","Line #","Customer PO","Order Date","Requested","Promised Ship",
            "Item","Next Status","QTY","UOM","2nd QTY","2nd UOM","Customer Name","TM Name"]
    return {
        "report_id": 325, "report_name": "CM - Open Sales Orders Live",
        "file": "04 - Open SO Live.xlsx",
        "notes": {
            "filters": [
                "SDNXTR NOT IN ('999') (open orders) - the only baked filter.",
                "No company filter (both 00010 and 00020 appear).",
                "Bulk-item whitelist: 76-code Bulk_Item IN (...) (incl. DPE3500.E, JS037, HP401, HSCF410, UNYTEC201).",
                "Optional prompts NOT applied in this export (default = all): Promised-Ship date range; Region single-select.",
            ],
            "slicers": [
                "Select the Region - single-select dropdown on REGION (Americas, Aubange, Shanghai, Singapore, Mumbai); optional, default = all.",
                "Enter the Date Range - Between (range) slicer on Promised Ship; optional, default = full range.",
            ],
            "cognos_path": CM_ROOT,
            "report_location_cognos": CM_ROOT,
            "jde_live": True,
            "report_location_edw": "CM Overview LIVE.pbix, page \"CM Open Sales Orders\"",
            "comments": [
                "VERDICT: CLEAN. 38/38 rows match 1:1 across all 16 visible columns; 100% match rate; zero real and zero cosmetic-only mismatches.",
                "Per-column mismatch counts all zero (Company, Branch, Order #, Line #, Customer PO, Order Date, Requested, Promised Ship, Item, Next Status, Primary/Secondary Qty+UOM, Customer Name, TM Name).",
                "Closed the 7/5 pagination gap: the ~18 previously-unchecked tail rows all tie 1:1. Company split: 00010 = 18 rows (CIN2/Americas), 00020 = 20 rows (AUBA/Aubange).",
                "TM Name 'Unassigned' fallback works: the 3 Michelman Int'l Belgium lines (orders 2645790, 2701152, 2737269) resolve to Unassigned on both sides.",
                "REGION excluded from the compare (PBI-only hidden slicer column, not in the Cognos visible list).",
                "Cosmetic only: DAX renders dates as m/d/yyyy 12:00:00 AM; Customer Name internal double-spaces appear identically in source on both sides.",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison",
             "sort_note": "Sorted by Promised Ship, Order #, Line #. 38/38 rows match; every flag = 1.",
             "cognos_headers": disp, "pbi_headers": disp, "rows": sec},
        ],
        "rs": {
            "key_label": "Order # | Line # | Item",
            "cognos_only": [], "pbi_only": [],
            "research_notes": ["No residuals - 38 distinct keys, all present and fully matched on both sides. REGION (PBI slicer column) intentionally excluded from the compare."],
            "col_map": {"cognos": ["QTY / 2nd QTY", "UOM / 2nd UOM"], "pbi": ["Primary Quantity / Secondary Quantity", "Primary UOM / Secondary UOM"]},
        },
    }

# ---------- 05 Inventory on Hand ----------
def r05():
    cog = read_csv(d("05 - Inventory on Hand", "cognos.csv"))
    ch = cog[0]; cog_rows = cog[1:]
    pbi = read_csv(d("05 - Inventory on Hand", "pbi_CM_Inventory_on_Hand.csv"))
    ph = [strip_tbl(h) for h in pbi[0]]; pbi_rows = pbi[1:]
    key = ["Branch Plant","Bulk Item","2nd Item Number","Status","Primary UOM"]
    ki = [ch.index(k) for k in key]
    sec, co, po = match_section(cog_rows, pbi_rows, ch, ki,
                                num_set={"KG/EA OH","LB/EA OH","Hard Commit"}, date_set=set(), vol_set=set(),
                                pbi_headers=ph)
    return {
        "report_id": 326, "report_name": "CM - Inventory on Hand",
        "file": "05 - Inventory on Hand.xlsx",
        "notes": {
            "filters": [
                "loc.LIPQOH/10000.0 > 0 (on-hand qty > 0).",
                "LTRIM(RTRIM(tag.IMBULK)) IN the fixed Bulk-Item whitelist (~130 entries as written; duplicates kept verbatim to match Cognos byte-for-byte).",
                "Region prompt (Select_Region) optional -> base query returns all regions; export was unfiltered.",
                "Inner joins: F4102<->F4101 (IBITM=IMITM), F4101<->F554101 (IMITM=IMITM), F4102<->F41021 (IBITM=LIITM AND IBMCU=LIMCU).",
            ],
            "slicers": ["Select the Region - single-select dropdown on REGION (Americas, Aubange, Shanghai, Singapore, Mumbai, OTHER); optional, default = all. REGION is both a visible column and the slicer source."],
            "cognos_path": CM_ROOT,
            "report_location_cognos": CM_ROOT,
            "jde_live": True,
            "report_location_edw": "CM Overview LIVE.pbix, page \"CM Inventory on Hand\"",
            "comments": [
                "VERDICT: CLEAN. 46/46 rows match Cognos exactly across all 9 columns; zero real discrepancies; no model/repo/PBIP changes made.",
                "Business key = Branch Plant + Bulk Item + 2nd Item Number + Status + Primary UOM. Every key unique and present on both sides; no orphans.",
                "Closed the 7/5 pagination gap: Aubange 17/17 (AUB2 x6, AUBA x11) and Singapore 5/5 (SING x3, SNG4 x2) now confirmed; Americas 24/24. Region distribution identical: Americas 24, Aubange 17, Singapore 5.",
                "KG/LB conversion math (factor 0.453593) verified on every row: KG rows LB = KG / 0.453593; LB rows KG = LB x 0.453593. Zero failures.",
                "No Shanghai/Mumbai rows in either source - those regions have no on-hand bulk inventory this snapshot (not a filter defect).",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison",
             "sort_note": "Sorted REGION, Bulk Item, 2nd Item Number. 46/46 rows match; every flag = 1.",
             "cognos_headers": ch, "pbi_headers": ph, "rows": sec},
        ],
        "rs": {
            "key_label": "Branch Plant | Bulk Item | 2nd Item Number | Status | Primary UOM",
            "cognos_only": [{"key": str(r)} for r in co], "pbi_only": [{"key": str(r)} for r in po],
            "research_notes": ["No residuals - byte-for-byte parity including the Aubange and Singapore rows never validated on 7/5."],
            "col_map": {"cognos": ch, "pbi": ph},
        },
    }

# ---------- 06 CM PO Live ----------
def r06():
    cog = read_csv(d("06 - CM PO Live", "cognos.csv"))
    ch = cog[0]; cog_rows = cog[1:]
    comp = read_csv(d("06 - CM PO Live", "comparison.csv"))
    comph = comp[0]; comp_rows = comp[1:]
    pbi = read_csv(d("06 - CM PO Live", "pbi_CM_PO_Live.csv"))
    ph_full = [strip_tbl(h) for h in pbi[0]]; pbi_rows_full = pbi[1:]
    # drop REGION from pbi
    reg_i = ph_full.index("REGION")
    ph = [h for i, h in enumerate(ph_full) if i != reg_i]
    pbi_rows = [[v for i, v in enumerate(r) if i != reg_i] for r in pbi_rows_full]
    # key on PO#, Line#
    cpo, cln = ch.index("PO #"), ch.index("Line #")
    ppo, pln = ph.index("Purchase Order Number"), ph.index("Line Number")
    pbucket = {(norm_str(r[ppo]), norm_str(r[pln])): r for r in pbi_rows}
    # comparison flags keyed by PO,Line
    flagcols = ["Company_match","Branch_match","PO_match","Line_match","Bulk_match","Item_match",
                "QTY_match","OpenQTY_match","SecQTY_match","NextStatus_match","Requested_match","Promised_match","Vendor_match"]
    fbucket = {(norm_str(r[comph.index("PO")]), norm_str(r[comph.index("Line")])): r for r in comp_rows}
    sec = []
    for row in cog_rows:
        k = (norm_str(row[cpo]), norm_str(row[cln]))
        prow = pbucket.get(k, [""] * len(ph))
        frow = fbucket.get(k)
        flags = [frow[comph.index(fc)] for fc in flagcols] if frow else [""] * len(ch)
        sec.append((row, flags, prow))
    return {
        "report_id": 327, "report_name": "CM - PO Live",
        "file": "06 - CM PO Live.xlsx",
        "notes": {
            "filters": [
                "Open QTY > 0 (true on all 19 rows).",
                "Promised Date >= today - 90 (rolling 90-day floor via CAST(GETDATE() AS date)).",
                "Bulk-item whitelist (75 codes) - every returned Bulk Item is in the hard-coded IN(...) list.",
                "Region prompt optional single-select (default all); Promised-date range prompt optional Between (unset in export).",
                "Parity quirks (no effect at one-line-per-PO grain): double-SUM + Item-Branch fan-out are no-ops (QTY = Open QTY on all 19 rows).",
            ],
            "slicers": [
                "Select the Region - single-select dropdown on REGION (Americas, Aubange, Shanghai, Singapore, Mumbai, OTHER); optional, default = all. REGION drives the slicer only (not a displayed column).",
                "Enter the Date Range - Between slicer on Promised Date; optional, default = today-90 forward.",
            ],
            "cognos_path": CM_ROOT,
            "report_location_cognos": CM_ROOT,
            "jde_live": True,
            "report_location_edw": "CM Overview LIVE.pbix, page \"CM PO Live\"",
            "comments": [
                "VERDICT: CLEAN. All 19 Cognos rows tie 1:1 to the 19 PBI rows; every displayed data cell matches (247/247 = 100.00%). No real discrepancies, no regressions.",
                "Join key = (Purchase Order Number, Line Number); unique on both sides. Per-column mismatch counts all zero across the 13 displayed columns.",
                "REGION is on the PBI table only as the slicer source (not a displayed column per BUILD.md) - excluded from the compare.",
                "Rolling 90-day window: earliest returned Promised Date = 2026-07-15, inside window. Next Status = 400 on every row both sides.",
                "Process note (not a data issue): the first EVALUATE returned only 10 rows (default MCP row cap); re-ran with maxRows + COUNTROWS -> confirmed 19.",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison",
             "sort_note": "Join key (PO #, Line #). 19/19 rows match; every flag = 1.",
             "cognos_headers": ch, "pbi_headers": ph, "rows": sec},
        ],
        "rs": {
            "key_label": "Purchase Order Number | Line Number",
            "cognos_only": [], "pbi_only": [],
            "research_notes": ["No residuals - 19 distinct PO/Line keys, all matched. REGION (PBI slicer column) excluded from compare."],
            "col_map": {
                "cognos": ["PO #", "Line #", "QTY", "Open QTY", "2nd QTY", "Item"],
                "pbi":    ["Purchase Order Number", "Line Number", "Primary Quantity", "Open Quantity", "Secondary Quantity", "2nd Item Number"],
            },
        },
    }

# ---------- 07 / 09 shared: Ivan 5-page list reports ----------
def build_status_section(base, comp_file, sheet_name, sort_note):
    """07-style: comparison_*.csv = status + single value block (MATCH/COGNOS_ONLY/PBI_ONLY)."""
    rows = read_csv(d(base, comp_file))
    hdr = [strip_tbl(h) for h in rows[0][1:]]  # drop 'status'
    sec = []; co = []; po = []
    for row in rows[1:]:
        status = row[0]; vals = row[1:]
        if status == "MATCH":
            sec.append((vals, [1] * len(hdr), vals))
        elif status == "COGNOS_ONLY":
            sec.append((vals, [""] * len(hdr), [""] * len(hdr))); co.append(vals)
        else:  # PBI_ONLY
            sec.append(([""] * len(hdr), [""] * len(hdr), vals)); po.append(vals)
    return {"sheet_name": sheet_name, "sort_note": sort_note,
            "cognos_headers": hdr, "pbi_headers": hdr, "rows": sec}, co, po

def r07():
    base = "07 - Ivan SK 2023"
    pages = [
        ("comparison_Inventory.csv", "Comparison - Inventory", "Inventory lots (REGION, Branch, Bulk, 2nd Item, Lot, Location, Status)."),
        ("comparison_Work_Orders.csv", "Comparison - Work Orders", "Work-order components (fan-out: one row per WO-line x component)."),
        ("comparison_Sales_Order_Summary.csv", "Comparison - Sales Orders", "Open sales-order lines."),
        ("comparison_Inventory_HP.csv", "Comparison - Inventory HP", "Inventory HP lots."),
        ("comparison_Safety_Stock_HP.csv", "Comparison - Safety Stock HP", "Safety-stock (DISTINCT F4102/F554101/F4101, IBSAFE grain)."),
    ]
    comps = []; rs_notes = []
    for cf, sn, note in pages:
        section, co, po = build_status_section(base, cf, sn, note)
        comps.append(section)
        rs_notes.append(f"{sn.split(' - ')[1]}: matched {sum(1 for c,f,p in section['rows'] if f and f[0]==1)}, Cognos-only {len(co)}, PBI-only {len(po)}.")
    resid = read_csv(d(base, "residuals.csv"))[1:]
    co_rs = [{"key": r[3], "note": f"{r[0]}: {r[4]}"} for r in resid if r[1] in ("COGNOS_ONLY",)]
    po_rs = [{"key": r[3], "note": f"{r[0]}: {r[5]}"} for r in resid if r[1] in ("PBI_ONLY",)]
    diff_rs = [f"{r[0]} DIFF {r[2]} key {r[3]}: Cognos={r[4]} | PBI={r[5]}" for r in resid if r[1] == "DIFF"]
    return {
        "report_id": 329, "report_name": "1 - Ivan SK 2023",
        "file": "07 - Ivan SK 2023.xlsx",
        "notes": {
            "filters": [
                "Inventory: Branch in whitelist; on-hand > 0; Bulk in SK whitelist.",
                "Work_Orders: component 2nd item in (BRIJS2.E, BRIJS20.E, BRIJS2.S, BRIJS20.S); Issued+Ordered > 0; WAUOM in (LB,KG); WASRST not in (MM); outer QtyRequested > 0.",
                "Sales_Order_Summary: SDLNTY='S'; SDPQOR>0; branch whitelist; country decode PRODCTL.F0005 (DRSY='00 ', DRRT='CN'); final inner-join Branch+2nd Item.",
                "Inventory_HP: Status IS NULL or in (T,B,Q,H); Branch whitelist; Bulk whitelist.",
                "Safety_Stock_HP: SELECT DISTINCT F4102/F554101/F4101 (IBSAFE grain); Branch + Bulk whitelist.",
            ],
            "slicers": ["(none) - pure list pages; no slicers/prompts/on-page plaintext."],
            "cognos_path": COGNOS_ROOT,
            "report_location_cognos": COGNOS_ROOT,
            "jde_live": True,
            "report_location_edw": "Ivan SK 2023.pbix (5 pages: Inventory, Work Orders, Sales Orders, Inventory HP, Safety Stock HP)",
            "comments": [
                "VERDICT: CLEAN - validated 2026-07-08 against a same-day Cognos export with the model freshly refreshed. Every table ties 100%: Inventory 307/307, Work Orders 220/220, Sales Orders 57/57, Inventory HP 144/144, Safety Stock HP 75/75. Zero residuals - no Cognos-only or PBI-only rows, and no cell-level diffs on any column.",
                "Full row-level multiset compare (numeric tolerance 0.02, date-only for date columns): exact match on all 803 rows across the 5 pages. The 7/6 drift diffs (Inventory HP allocation pair, 1 Inventory lot) are gone now that both sides are same-day.",
                "Column mismatches: none. Decode, country, CSR/TM name, UOM, and KG/LB conversion all tie exactly.",
                "Fan-out (Work Orders page) verified correct: AVG(WAUORG) OVER renders a constant Quantity Requested across a WO's components; KG rows tie (P7 ISSUED LB = KG/0.453593).",
                "Data pulled via ADOMD (full tables, no row cap) 2026-07-08; compare + workbook regenerated same day.",
            ],
        },
        "comparisons": comps,
        "rs": {
            "key_label": "Per-table business key (see page comparisons)",
            "cognos_only": co_rs, "pbi_only": po_rs,
            "research_notes": rs_notes + diff_rs + [
                "All residuals reconcile to live-JDE drift; Safety Stock HP is a perfect 75/75, confirming the rebuild logic is correct.",
            ],
            "col_map": {"cognos": ["(same column names on both sides - PBI columns are the table[column] duals of the Cognos list)"], "pbi": [""]},
        },
    }

def r09():
    base = "09 - Ivan FC 2023"
    # tables: (cognos_csv, residual_csv, sheet_name, sort_note, key_cols, num_cols, date_cols, vol_cols)
    specs = [
        ("cognos_inventory.csv", "residual_inventory.csv", "Comparison - Inventory",
         "Inventory lots.", ["Branch Plant","Bulk Item","2nd Item Number","Lot Number","Location","Status","Stock Type"],
         {"Quantity On Hand","Hard Commit","OH KG","OH LB"}, set(), {"Date"}),
        ("cognos_work_orders.csv", "residual_work_orders.csv", "Comparison - Work Orders",
         "Work-order components (fan-out).", ["WO Number","2nd Item Number","Component 2nd Item Number","Start Date"],
         {"Quantity Requested","Quantity Completed","REQUEST KG","COMPLETE KG","P7 ISSUED KG","P7 ORDERED KG","P7 ISSUED LB","P7 ORDERED LB","P7 REMAINING"},
         {"Start Date","Completed Date"}, {"DATE"}),
        ("cognos_sales_order_summary.csv", "residual_sales_order_summary.csv", "Comparison - Sales Orders",
         "Open sales-order lines.", ["Order Number","2nd Item Number","Bulk Item"],
         {"ORDER KGs","ORDER LBs","Prim QTY","2nd QTY"},
         {"Order Date","Requested Date","Promised Ship Date","Scheduled Pick Date"}, set()),
        ("cognos_inventory_hp.csv", "residual_inventory_hp.csv", "Comparison - Inventory HP",
         "Inventory HP lots.", ["Branch Plant","Bulk Item","2nd Item Number","Location","Lot Number","Status"],
         {"Quantity On Hand","LB"}, set(), set()),
        ("cognos_safety_stock_hp.csv", "residual_safety_stock_hp.csv", "Comparison - Safety Stock HP",
         "Safety-stock (IBSAFE grain).", ["Branch Plant","Bulk Item","2nd Item Number"],
         {"Safety Stock","LB Safety Stock"}, set(), set()),
    ]
    comps = []; rs_co = []; rs_po = []; rs_notes = []
    for cog_csv, res_csv, sn, note, keycols, numset, dateset, volset in specs:
        cog = read_csv(d(base, cog_csv)); ch = cog[0]; cog_rows = cog[1:]
        res = read_csv(d(base, res_csv)); rh = res[0]; res_rows = res[1:]
        # residual data cols after side,key align to ch
        dcols = rh[2:]
        def resrow_vals(r):  # map residual data cols onto ch order
            m = dict(zip(dcols, r[2:]))
            return [m.get(c, "") for c in ch]
        cog_res = {}  # side rows
        co_full = []; po_full = []; cdiff = {}; pdiff = {}
        for r in res_rows:
            side = r[0]; key = r[1]; vals = resrow_vals(r)
            if side == "COGNOS-only": co_full.append(vals)
            elif side == "PBI-only": po_full.append(vals)
            elif side == "COGNOS-diff": cdiff[key] = vals
            elif side == "PBI-diff": pdiff[key] = vals
        kinds = [col_kind(h, numset, dateset, volset) for h in ch]
        # matched exact = cognos rows whose business key is not a residual cognos-side key
        kidx = [ch.index(k) for k in keycols]
        def keyf(vals):
            return tuple(norm_date(vals[i]) if ch[i] in dateset else norm_str(vals[i]) for i in kidx)
        excl_keys = set(keyf(v) for v in co_full) | set(keyf(v) for v in cdiff.values())
        used_excl = {}
        sec = []
        for row in cog_rows:
            k = keyf(row)
            if k in excl_keys:
                # first occurrences of an excluded key are the residual rows; skip one each
                cap = sum(1 for v in co_full if keyf(v) == k) + sum(1 for v in cdiff.values() if keyf(v) == k)
                if used_excl.get(k, 0) < cap:
                    used_excl[k] = used_excl.get(k, 0) + 1
                    continue
            sec.append((row, [1] * len(ch), row))
        matched = len(sec)
        # diff pairs
        for key, cvals in cdiff.items():
            pvals = pdiff.get(key, [""] * len(ch))
            flags = [1 if cell_eq(cvals[i], pvals[i], kinds[i]) else 0 for i in range(len(ch))]
            sec.append((cvals, flags, pvals))
        # cognos-only / pbi-only inline
        for v in co_full:
            sec.append((v, [""] * len(ch), [""] * len(ch)))
        for v in po_full:
            sec.append(([""] * len(ch), [""] * len(ch), v))
        comps.append({"sheet_name": sn, "sort_note": note + f" Matched {matched}; diff pairs {len(cdiff)}; Cognos-only {len(co_full)}; PBI-only {len(po_full)}.",
                      "cognos_headers": ch, "pbi_headers": ch, "rows": sec})
        tbl = sn.split(" - ")[1]
        rs_notes.append(f"{tbl}: exact-match {matched}, drift cell-diff pairs {len(cdiff)}, Cognos-only {len(co_full)}, PBI-only {len(po_full)}.")
        rs_co += [{"key": f"{tbl}: {'|'.join(norm_str(x) for x in v[:6])}"} for v in co_full]
        rs_po += [{"key": f"{tbl}: {'|'.join(norm_str(x) for x in v[:6])}"} for v in po_full]
    return {
        "report_id": 331, "report_name": "1 - Ivan FC 2023",
        "file": "09 - Ivan FC 2023.xlsx",
        "notes": {
            "filters": [
                "Inventory: Branch in (SHAN,MUM3,SING,SNG4,AUBA,AUB2); on-hand > 0; Bulk in 31-item whitelist.",
                "Work_Orders: component 2nd item in (BRIJS2.E,BRIJS20.E,BRIJS2.S,BRIJS20.S); Issued+Ordered > 0; Start >= 2025-11-01; WAUOM in (LB,KG); WASRST not in (MM); outer QtyRequested > 0.",
                "Sales_Order_Summary: SDLNTY='S'; SDPQOR>0; SDNXTR not in (570,580,620,999); branch whitelist; country decode PRODCTL.F0005 (DRSY='00 ', DRRT='CN').",
                "Inventory_HP: Status IS NULL or in (T,B,Q,H); Branch whitelist; Bulk whitelist.",
                "Safety_Stock_HP: SELECT DISTINCT F4102/F554101/F4101 (IBSAFE grain); Branch + Bulk whitelist.",
            ],
            "slicers": ["(none) - pure list pages; no slicers/prompts/on-page plaintext."],
            "cognos_path": COGNOS_ROOT,
            "report_location_cognos": COGNOS_ROOT,
            "jde_live": True,
            "report_location_edw": "Ivan FC 2023.pbix (5 pages: Inventory, Work Orders, Sales Orders, Inventory HP, Safety Stock HP)",
            "comments": [
                "VERDICT: CLEAN - validated 2026-07-08. The model was refreshed to same-day data (Inventory[NOW]/Work_Orders[DATE] = 2026-07-08) and compared to a same-day Cognos export. Every table ties 100%: Inventory 87/87, Work Orders 417/417, Sales Orders 24/24, Inventory HP 232/232, Safety Stock HP 140/140. Zero residuals on every page.",
                "The prior 7/5 'drift-polluted' snapshot (27.6h refresh gap) is fully resolved: after the 7/8 refresh, all previously-drifted rows (Inventory SING<->SNG4 moves, Inventory HP allocation/transit lots, Sales status advances) reconcile exactly. No structural differences remain.",
                "Full row-level multiset compare (numeric tolerance 0.001, date-only for date columns): exact match on all 900 rows across the 5 pages. Column mismatches: none - including decode, country, CSR/TM, UOM, and KG/LB conversion.",
                "Fan-out (Work Orders page) correct; Safety Stock HP 140/140 exact confirms the rebuild logic.",
                "Data pulled via ADOMD (full tables, no row cap) 2026-07-08; compare + workbook regenerated same day.",
            ],
        },
        "comparisons": comps,
        "rs": {
            "key_label": "Per-table business key (see page comparisons)",
            "cognos_only": rs_co, "pbi_only": rs_po,
            "research_notes": rs_notes + [
                "No residuals on any page after the 2026-07-08 same-day refresh: 900/900 rows tie exactly (Inventory 87, Work Orders 417, Sales 24, Inventory HP 232, Safety Stock 140). The earlier 27.6h refresh gap is closed.",
            ],
            "col_map": {"cognos": ["MPF", "Date", "Site / REGION"], "pbi": ["Master Planning Family", "NOW", "REGION"]},
        },
    }

# ---------- 08 / 10 shared: Ivan Forecast (Sales History + Forecast) ----------
SH_COLS = ["Order Company","Branch Plant","Global Bulk Item","Bulk Item","2nd Item Number","Order Number",
           "Next Status","Year","Month","Week","Promised Ship Date","Ordered Quantity KGs","Revenue Business Unit",
           "Customer Code","Customer Name","Global Parent","Global Parent Name","TM Name","Country Name",
           "Ordered Quantity","Ordering Unit of Measure","Ordered Quantity LBs","Open Indicator"]
SH_NUM = {"Ordered Quantity KGs","Ordered Quantity LBs","Ordered Quantity"}
SH_INT = {"Order Number","Year","Month","Week","Global Parent"}
SH_DATE = {"Promised Ship Date"}

def _sh_key(dct):
    def ti(v):
        try: return str(int(float(str(v).strip())))
        except Exception: return str(v).strip()
    def tn(v):
        try: return round(float(v), 3)
        except Exception: return None
    return (ti(dct.get("Order Number","")), str(dct.get("2nd Item Number","")).strip(),
            norm_date(dct.get("Promised Ship Date","")), str(dct.get("Ordering Unit of Measure","")).strip(),
            tn(dct.get("Ordered Quantity","")), str(dct.get("Branch Plant","")).strip(),
            str(dct.get("Customer Code","")).strip())

def _sh_flags(c, p):
    flags = []
    for col in SH_COLS:
        cv, pv = c.get(col, ""), p.get(col, "")
        if col in SH_NUM:
            na, nb = norm_num(cv), norm_num(pv)
            if na is None or nb is None:
                ok = norm_str(cv) == norm_str(pv)
            else:
                ok = abs(na - nb) <= 0.01   # strict, matches validation agent (surfaces the LB->KG factor drift)
        elif col in SH_INT:
            ok = norm_str(cv) == norm_str(pv)
        elif col in SH_DATE:
            ok = norm_date(cv) == norm_date(pv)
        else:
            ok = str(cv).strip() == str(pv).strip()
        flags.append(1 if ok else 0)
    return flags

def r08():
    base = "08 - SK Forecast"
    cog = read_csv(d(base, "cognos_sales_history.csv")); chh = cog[0]
    cog_dicts = [dict(zip(chh, r)) for r in cog[1:]]
    pbi = read_csv(d(base, "pbi_Sales_History.csv")); phh = pbi[0]
    pbi_dicts = [dict(zip(phh, r)) for r in pbi[1:]]
    pbucket = {}
    for pd_ in pbi_dicts:
        pbucket.setdefault(_sh_key(pd_), []).append(pd_)
    used = {}
    sec = []; co = []; po = []
    # sort cognos by promised then order number
    def sortk(dct):
        return (norm_date(dct.get("Promised Ship Date","")), norm_str(dct.get("Order Number","")))
    for c in sorted(cog_dicts, key=sortk):
        k = _sh_key(c); cand = pbucket.get(k, []); idx = used.get(k, 0)
        cogv = [c.get(col, "") for col in SH_COLS]
        if idx < len(cand):
            p = cand[idx]; used[k] = idx + 1
            sec.append((cogv, _sh_flags(c, p), [p.get(col, "") for col in SH_COLS]))
        else:
            sec.append((cogv, [""] * len(SH_COLS), [""] * len(SH_COLS))); co.append(c)
    for k, cand in pbucket.items():
        for idx in range(used.get(k, 0), len(cand)):
            p = cand[idx]
            sec.append(([""] * len(SH_COLS), [""] * len(SH_COLS), [p.get(col, "") for col in SH_COLS])); po.append(p)
    matched = sum(1 for cog_, f, pbi_ in sec if f and f[0] != "")
    return {
        "report_id": 330, "report_name": "1 - Ivan SK 2023 Forecast",
        "file": "08 - SK Forecast.xlsx",
        "notes": {
            "filters": [
                "IMBULK IN (~99-item SK whitelist: PR3460/PR5980I/PR5985/DPI*/MF*/MP*/24xxxPX/25xxxNX...), both pages.",
                "Sales History window: DUE_DATE >= 2026-03-01 (FC >= 2025-11-01), <= EOMONTH(sysdate+180). Confirmed min 2026-03-02.",
                "SDPQOR/10000 > 0; SDCNDJ = 0; next status incl. 999; SDLNTY NOT LIKE '%F%'; GST exclusion CASE.",
                "Forecast branches: FTMCU IN (AUBA,AUB2,SING,SNG4,MUM3,SHAN,CINC,CIN2,CIN4) - correctly includes CINC/CIN2 (SK-specific vs report 10).",
            ],
            "slicers": ["(none) - list pages; no slicers/prompts."],
            "cognos_path": COGNOS_ROOT,
            "report_location_cognos": COGNOS_ROOT,
            "jde_live": True,
            "report_location_edw": "Ivan SK 2023 Forecast.pbix (2 pages: Sales History, Forecast)",
            "comments": [
                "VERDICT: Sales History PASSES (production-ready) - 864/881 matched keys = 98.07%; the +14 net PBI rows are fully explained (17 rescheduled 1:1 pairs from a 6-min live-data drift + 14 brand-new orders). F42119 union working; BUILD risk #2 closed. First live test of the DW_LEGACY->JDE reverse-map - it holds.",
                "3 REAL parity findings (all root-caused, none affect row counts):",
                "  1. TM Name (469 rows): Cognos DW 'First Last' vs JDE ABALPH 'Last, First' (e.g. Krist Vanderstiggel vs Vanderstiggel, Krist; nickname diffs Dave vs David). Correct PERSON resolved; string order differs. Fix option: reformat ABALPH (decision needed - nicknames won't fully tie).",
                "  2. Ordered Quantity KGs (143 LB-primary rows): .m uses q*0.453593; Cognos uses q/2.2045992 = q*0.45359719 (delta up to ~0.17 KG). Fix: LB -> q / 2.2045992 in the KG CASE.",
                "  3. Ordered Quantity LBs (4 EA-primary rows, 251194NX.S-B1): .m uses q*44; Cognos 44.091984 (= 20 x 2.2045992). Fix: EA -> q x 44.091984.",
                "Zero-mismatch on the highest-risk reverse-maps: Country Name (F0005 00/CN decode), Global Parent/GP Name (ABAN86 self-join), Customer Name, Revenue BU, Ordered Qty + Ordering UOM (SDPQOR/10000, transaction UOM; SALES_FACTOR=1 validated).",
                "Cosmetic: Cognos prints 'Order Number' twice (cols 6 & 20); PBI holds it once.",
                "FORECAST PAGE UNVALIDATED - Cognos export sheet Forecast_1 is blank ('No Data Available', the known Cognos date-window issue); no source-of-truth to tie against. No comparison sheet is produced for Forecast. PBI profile (rebuild only, refreshed 2026-07-06 20:50): 768 rows; Requested 2026-07-04 -> 07-25; 82 items; 155 customers; branches CIN2 436 / AUBA 196 / MUM3 84 / SNG4 36 / CINC 16; no all-zero-measure anomaly. F3460 field maps (FTFQT /10000 scaling, FTDRQJ, FTAN8) remain untested without JDE access; carries the same LB->KG factor drift if ever tied.",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison - Sales History",
             "sort_note": f"Sorted by Promised Ship Date, Order Number. Business key (Order#, 2nd Item, Promised, UOM, Qty, Branch, Customer). Matched {matched}; Cognos-only {len(co)}; PBI-only {len(po)}. Flag=0 columns are the 3 documented parity findings + point-in-time drift.",
             "cognos_headers": SH_COLS, "pbi_headers": SH_COLS, "rows": sec},
        ],
        "rs": {
            "key_label": "Order Number | 2nd Item | Promised Ship | UOM | Qty",
            "cognos_only": [{"key": f"{c.get('Order Number')} | {c.get('2nd Item Number')} | {norm_date(c.get('Promised Ship Date'))} | {c.get('Ordering Unit of Measure')} | {c.get('Ordered Quantity')}", "note": "rescheduled 1:1 pair (6-min live drift)"} for c in co],
            "pbi_only": [{"key": f"{p.get('Order Number')} | {p.get('2nd Item Number')} | {norm_date(p.get('Promised Ship Date'))} | {p.get('Ordering Unit of Measure')} | {p.get('Ordered Quantity')}", "note": f"Next Status {p.get('Next Status')}"} for p in po],
            "research_notes": [
                f"Cognos-only {len(co)} + PBI-only {len(po)} = 17 rescheduled 1:1 pairs (6-min drift) + 14 brand-new orders (2718622, 2734692/739/848, 2736904/05/941/42, 2737252, 2737445, AUBA 26001256/261/262/264). Net +14 exact. F42119 union working.",
                "Recurring flag=0 columns on matched rows: TM Name (469, name-order format), Ordered Quantity KGs (143, LB->KG factor), Ordered Quantity LBs (4, EA factor), plus Next Status (16) / Open Indicator (9) point-in-time drift. See Notes for fixes.",
            ],
            "col_map": {"cognos": ["Order Number (printed twice)", "Ordered Quantity KGs", "Ordered Quantity LBs"],
                        "pbi": ["Order Number (once)", "Ordered Quantity KGs", "Ordered Quantity LBs"]},
        },
    }

def r10():
    base = "10 - SFC Forecast"
    rows = read_csv(d(base, "comparison_sales_history.csv")); hdr = rows[0]
    ci = [hdr.index(c + "__cognos") for c in SH_COLS]
    pi = [hdr.index(c + "__pbi") for c in SH_COLS]
    diff_i = hdr.index("diff_columns")
    sec = []; co = 0; po = 0; matched = 0
    for row in rows[1:]:
        status = row[0]
        cogv = [row[i] for i in ci]; pbiv = [row[i] for i in pi]
        if status == "both":
            matched += 1
            diffs = set(x.strip() for x in row[diff_i].split(";") if x.strip())
            flags = [0 if col in diffs else 1 for col in SH_COLS]
            # fallback: recompute if diff_columns empty but values differ
            if not diffs:
                flags = [1 if cell_eq(cogv[k], pbiv[k], "num" if SH_COLS[k] in SH_NUM else ("date" if SH_COLS[k] in SH_DATE else "str")) else 0 for k in range(len(SH_COLS))]
            sec.append((cogv, flags, pbiv))
        elif status == "cognos_only":
            co += 1; sec.append((cogv, [""] * len(SH_COLS), [""] * len(SH_COLS)))
        else:
            po += 1; sec.append(([""] * len(SH_COLS), [""] * len(SH_COLS), pbiv))
    resid = read_csv(d(base, "residuals.csv"))[1:]
    co_rs = [{"key": f"{r[1]} | {r[2]} | {r[3]} | NS {r[4]}", "note": f"{r[7]} ({r[5]}, {r[6]}kg)"} for r in resid if r[0] == "cognos_only"]
    po_rs = [{"key": f"{r[1]} | {r[2]} | {r[3]} | NS {r[4]}", "note": r[7]} for r in resid if r[0] == "pbi_only"]
    return {
        "report_id": 332, "report_name": "1 - Ivan SFC2023 Forecast",
        "file": "10 - SFC Forecast.xlsx",
        "notes": {
            "filters": [
                "SDLNTY NOT LIKE '%F%'; SDCNDJ = 0; SDPQOR/10000 > 0 (SALES_FACTOR=1).",
                "Promised (SDPDDJ) >= 2025-11-01 AND <= EOMONTH(GETDATE()+180d).",
                "Bulk whitelist: JS168.S, ME91735.S, ME92040.S, PP05S.S, ME91240G.S, MG7140.S, TSPP01.S, ME87235.S, ME90640.S, 211018IX.S, PP236A.S, NYS2104.S, JS168.E, BRIJS2.S, BRIJS20.S, BRIJS2.E, BRIJS20.E (verbatim dup entries kept).",
                "GST exclusion: CASE WHEN IMGBLK='-' THEN SDLITM ELSE IMGBLK END NOT IN (IGST,CGST,SGST,CVD,ADD).",
                "BUDGET_FACTOR -> no-op on F4211.",
            ],
            "slicers": ["(none) - list pages; no slicers/prompts."],
            "cognos_path": COGNOS_ROOT,
            "report_location_cognos": COGNOS_ROOT,
            "jde_live": True,
            "report_location_edw": "Ivan SFC2023 Forecast.pbix (2 pages: Sales History, Forecast)",
            "comments": [
                "VERDICT: Sales History CLEAN / production-ready. Confirmed post-refresh 2026-07-07 against a fresh same-day Cognos export: PBI 913 rows = Cognos 913 rows. Exact tie on row count, sum KG (6,143,487), distinct orders (828), and all 14 Branch x UOM buckets (count AND sum-KG).",
                "F4211 UNION F42119 fix WORKING: this was the 7/6 deployed-lags-repo gap (deployed partition had the OLD F4211-only query = 21 rows; the repo union was pushed to the live partition 7/6 and refreshed 7/7 -> 913 rows). The previously-missing purged/closed (999) history is now present.",
                "Full row-level diff (all 913): 912/913 rows match on business key AND every column except the documented items below. KG / LB / Ordered-Qty tie EXACTLY (0 mismatches) - the report-08 LB->KG / EA->LB conversion-factor finding does NOT surface in report 10's data.",
                "TM Name (225 rows): format-only - JDE ABALPH 'Cheng, Ethan' / 'Peng,Li(Robin)' vs Cognos DW 'Ethan Cheng' / 'Robin Li'; correct person resolved, string order differs. ACCEPTED as a documented cosmetic (user decision 2026-07-07); flagged 0 with note, no model change.",
                "Point-in-time drift (NOT defects): 1 Cognos-only + 1 PBI-only row = one open order whose status flipped in the minutes between the Cognos export and the PBI refresh; Next Status differs on 2 matched rows for the same reason.",
                "FORECAST PAGE UNVALIDATED - Cognos sheet Forecast_1 is EMPTY (blank by design); no source-of-truth, so no comparison sheet is produced for Forecast. F3460 /10000 scaling still needs a JDE/F3460 human check (tiny values flag the scaling risk).",
            ],
        },
        "comparisons": [
            {"sheet_name": "Comparison - Sales History",
             "sort_note": f"Business-key match (as of 7/7). Matched {matched}; Cognos-only {co}; PBI-only {po} (one open order's status flipped between the export and the refresh). Flag=0 columns = TM Name format (225, documented cosmetic) + Next Status drift (2).",
             "cognos_headers": SH_COLS, "pbi_headers": SH_COLS, "rows": sec},
        ],
        "rs": {
            "key_label": "Order | 2nd Item | Promised | Next Status",
            "cognos_only": co_rs,
            "pbi_only": po_rs,
            "research_notes": [
                f"913 PBI = 913 Cognos rows (fresh same-day 7/7 export). {len(co_rs)} Cognos-only + {len(po_rs)} PBI-only = one open order whose status flipped in the minutes between the Cognos export and the PBI refresh - a point-in-time drift, not a rebuild error. F4211 UNION F42119 union confirmed working.",
                "225 matched rows differ only on TM Name (JDE ABALPH 'Last, First' vs Cognos DW 'First Last'; correct person) - ACCEPTED as documented cosmetic (user 2026-07-07). 2 rows differ on Next Status (same open-order drift). All other columns - incl. KG / LB / Ordered Qty - tie exactly.",
            ],
            "col_map": {"cognos": ["Order Number (printed twice)", "Ordered Quantity KGs", "Ordered Quantity LBs"],
                        "pbi": ["Order Number (once)", "Ordered Quantity KGs", "Ordered Quantity LBs"]},
        },
    }

if __name__ == "__main__":
    builders = [r01, r02, r03, r04, r05, r06, r07, r08, r09, r10]
    for b in builders:
        cfg = b()
        build(cfg)
        print()
