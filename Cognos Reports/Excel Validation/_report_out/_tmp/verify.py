# -*- coding: utf-8 -*-
"""Reopen each finished workbook and assert structure + row counts."""
import os
from openpyxl import load_workbook

OUT = r"C:\Users\Zack\Documents\Code\michelman\Cognos Reports\Excel Validation\_report_out"

EXPECT = {
    "01 - RM Staging.xlsx": {"sheets": ["Notes","Comparison - RM Requirements","Comparison - Work Order Detail","RS"],
                             "rows": {"Comparison - RM Requirements":8, "Comparison - Work Order Detail":15}},
    "02 - 530 Report.xlsx": {"sheets": ["Notes","Comparison","RS"], "rows": {"Comparison":43}},
    "03 - SO under 560.xlsx": {"sheets": ["Notes","Comparison - Sales Orders","Comparison - Inventory","Comparison - Subtotals","RS"],
                               "rows": {"Comparison - Sales Orders":2,"Comparison - Inventory":2,"Comparison - Subtotals":2}},
    "04 - Open SO Live.xlsx": {"sheets": ["Notes","Comparison","RS"], "rows": {"Comparison":38}},
    "05 - Inventory on Hand.xlsx": {"sheets": ["Notes","Comparison","RS"], "rows": {"Comparison":46}},
    "06 - CM PO Live.xlsx": {"sheets": ["Notes","Comparison","RS"], "rows": {"Comparison":19}},
    "07 - Ivan SK 2023.xlsx": {"sheets": ["Notes","Comparison - Inventory","Comparison - Work Orders","Comparison - Sales Orders","Comparison - Inventory HP","Comparison - Safety Stock HP","RS"],
                               "rows": {"Comparison - Inventory":308,"Comparison - Work Orders":216,"Comparison - Sales Orders":52,"Comparison - Inventory HP":135,"Comparison - Safety Stock HP":75}},
    "08 - SK Forecast.xlsx": {"sheets": ["Notes","Comparison - Sales History","RS"], "rows": {"Comparison - Sales History":912}},
    "09 - Ivan FC 2023.xlsx": {"sheets": ["Notes","Comparison - Inventory","Comparison - Work Orders","Comparison - Sales Orders","Comparison - Inventory HP","Comparison - Safety Stock HP","RS"],
                               "rows": {"Comparison - Inventory":87,"Comparison - Sales Orders":25,"Comparison - Inventory HP":250,"Comparison - Safety Stock HP":140}},
    "10 - SFC Forecast.xlsx": {"sheets": ["Notes","Comparison - Sales History","RS"], "rows": {"Comparison - Sales History":909}},
}

def data_rows(ws):
    # data starts row 10; count rows with any value in col A..end
    n = 0
    for r in range(10, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            n = r
    return (n - 9) if n >= 10 else 0

ok = True
for fn, exp in EXPECT.items():
    p = os.path.join(OUT, fn)
    wb = load_workbook(p)
    names = wb.sheetnames
    problems = []
    if names != exp["sheets"]:
        problems.append(f"SHEETS {names} != {exp['sheets']}")
    # Notes title check
    notes = wb["Notes"]
    if not str(notes["A1"].value or "").startswith("Cognos Migration - Report ID"):
        problems.append(f"Notes A1 bad: {notes['A1'].value!r}")
    if str(notes["A2"].value) != "Data revised 7/6/2026":
        problems.append(f"Notes A2 bad: {notes['A2'].value!r}")
    if notes["A1"].font.name != "Tahoma" or notes["A1"].font.size != 14 or not notes["A1"].font.bold:
        problems.append("Notes A1 font not Tahoma 14 bold")
    # comparison sheets
    for sh in names:
        if not sh.startswith("Comparison"):
            continue
        ws = wb[sh]
        if str(ws["A1"].value) != "Cognos Report":
            problems.append(f"{sh} A1 != 'Cognos Report'")
        if str(ws["A2"].value) != "As of 7/6/2026":
            problems.append(f"{sh} A2 bad")
        # row 7 labels
        labels = [ws.cell(7, c).value for c in range(1, ws.max_column + 1) if ws.cell(7, c).value]
        if not ({"Cognos","Compare","PBI"} <= set(labels)):
            problems.append(f"{sh} row7 labels missing: {labels}")
        # header row 9 has content
        if not any(ws.cell(9, c).value for c in range(1, ws.max_column + 1)):
            problems.append(f"{sh} row9 headers empty")
        # freeze
        if ws.freeze_panes != "A10":
            problems.append(f"{sh} freeze={ws.freeze_panes}")
        dr = data_rows(ws)
        if sh in exp["rows"] and dr != exp["rows"][sh]:
            problems.append(f"{sh} data rows {dr} != {exp['rows'][sh]}")
    # RS present
    if "RS" not in names:
        problems.append("no RS")
    # font check: sample a comparison data cell
    status = "OK" if not problems else "FAIL"
    if problems: ok = False
    print(f"[{status}] {fn}")
    for pr in problems:
        print("      -", pr)

print("\nALL GOOD" if ok else "\nISSUES FOUND")
