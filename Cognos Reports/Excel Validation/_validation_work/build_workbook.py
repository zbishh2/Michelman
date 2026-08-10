# ============================================================================
# Report-out workbook builder - reproduces the "Missing COA Rohit.xlsx" format:
#   Sheet 1 "Notes"       - title, data-revised date, Cognos filters, slicers,
#                           links, report locations, perspective, comments
#   Sheet 2+ "Comparison" - Cognos block | per-column 1/0 Compare block | PBI
#                           block, section labels on row 7, headers row 9,
#                           data from row 10 (one comparison sheet per section
#                           for multi-page reports)
#   Last sheet "RS"       - residual keys (in Cognos but not PBI / vice versa)
#                           with research notes + Cognos<->PBI column-name map
#
# Usage: from build_workbook import build_workbook; build_workbook(config, out)
# config schema documented at bottom. Data comes from each report's
# _validation_work\<NN>\ CSVs produced by the validation agents.
# ============================================================================
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

TAHOMA = "Tahoma"
F_BASE = Font(name=TAHOMA, size=10)
F_BOLD = Font(name=TAHOMA, size=10, bold=True)
F_TITLE = Font(name=TAHOMA, size=14, bold=True)
FILL_HDR = PatternFill("solid", fgColor="FFE7E5E5")
FILL_SECT = PatternFill("solid", fgColor="FFF2F2F2")
FILL_BAD = PatternFill("solid", fgColor="FFFFC7CE")   # red-ish for 0 flags
THIN = Side(style="thin", color="FFBFBFBF")
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return rows


def _autofit(ws, min_w=8, max_w=42):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                w = min(max(len(str(c.value)) + 2, min_w), max_w)
                col = c.column_letter
                if w > widths.get(col, 0):
                    widths[col] = w
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _notes_sheet(ws, cfg):
    ws.sheet_view.showGridLines = False
    n = cfg["notes"]
    ws["A1"] = f"Cognos Migration - Report ID {cfg['report_id']} - {cfg['report_name']}"
    ws["A1"].font = F_TITLE
    ws["A2"] = f"Data revised {cfg['data_revised']}"
    ws["A2"].font = F_BASE
    r = 6
    ws.cell(r, 3, "Cognos filters").font = F_BOLD
    r += 1
    for line in n.get("filters", []):
        ws.cell(r, 3, line).font = F_BASE
        r += 1
    r += 1
    ws.cell(r, 3, "Slicers").font = F_BOLD
    r += 1
    for line in n.get("slicers", []):
        ws.cell(r, 3, line).font = F_BASE
        r += 1
    r += 1
    ws.cell(r, 3, "Links").font = F_BOLD
    r += 1
    links = n.get("links", {})
    for label in ("Power BI", "Cognos"):
        cell = ws.cell(r, 4, label)
        cell.font = F_BASE
        url = links.get(label.lower().replace(" ", ""))
        if url:
            cell.hyperlink = url
            cell.font = Font(name=TAHOMA, size=10, color="FF0563C1", underline="single")
        r += 1
    r += 1
    ws.cell(r, 3, "Report Location").font = F_BOLD
    r += 1
    ws.cell(r, 3, "  Cognos").font = F_BASE
    ws.cell(r, 4, n.get("report_location_cognos", "")).font = F_BASE
    r += 1
    ws.cell(r, 4, "  Note report uses JDE live data" if n.get("jde_live") else "").font = F_BASE
    r += 1
    ws.cell(r, 3, "  EDW").font = F_BASE
    ws.cell(r, 4, n.get("report_location_edw", "")).font = F_BASE
    r += 2
    if n.get("perspective"):
        ws.cell(r, 3, "  EDW Perspective Used").font = F_BASE
        r += 1
        ws.cell(r, 4, n["perspective"]).font = F_BASE
        r += 2
    ws.cell(r, 3, "  Comments").font = F_BOLD
    r += 1
    for block in n.get("comments", []):
        for line in str(block).splitlines():
            ws.cell(r, 3, line).font = F_BASE
            r += 1
        r += 1
    _autofit(ws)


def _comparison_sheet(ws, cfg, section):
    """section: {title, sort_note, cognos_headers, pbi_headers, rows}
    rows: list of (cognos_vals, flag_vals, pbi_vals) value-tuples."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Cognos Report"
    ws["A1"].font = F_BOLD
    ws["A2"] = f"As of {cfg['data_revised']}"
    ws["A2"].font = F_BASE
    if section.get("sort_note"):
        ws["A4"] = section["sort_note"]
        ws["A4"].font = F_BASE
    ch = section["cognos_headers"]
    ph = section["pbi_headers"]
    nC, nP = len(ch), len(ph)
    c0 = 1                     # Cognos block start col (A)
    j0 = c0 + nC + 1           # Compare block start (1 sep col)
    s0 = j0 + nC + 1           # PBI block start (1 sep col)
    # row 7 section labels
    for col, label in ((c0, "Cognos"), (j0, "Compare"), (s0, "PBI")):
        cell = ws.cell(7, col, label)
        cell.font = F_BOLD
        cell.fill = FILL_SECT
    # row 9 headers
    for i, h in enumerate(ch):
        cell = ws.cell(9, c0 + i, h)
        cell.font = F_BOLD; cell.fill = FILL_HDR; cell.border = B_ALL
    for i, h in enumerate(ch):
        cell = ws.cell(9, j0 + i, h)
        cell.font = F_BOLD; cell.fill = FILL_HDR; cell.border = B_ALL
    for i, h in enumerate(ph):
        cell = ws.cell(9, s0 + i, h)
        cell.font = F_BOLD; cell.fill = FILL_HDR; cell.border = B_ALL
    # data from row 10
    r = 10
    for cog, flags, pbi in section["rows"]:
        for i, v in enumerate(cog):
            ws.cell(r, c0 + i, v).font = F_BASE
        for i, v in enumerate(flags):
            cell = ws.cell(r, j0 + i, v)
            cell.font = F_BASE
            if str(v) == "0":
                cell.fill = FILL_BAD
        for i, v in enumerate(pbi):
            ws.cell(r, s0 + i, v).font = F_BASE
        r += 1
    ws.freeze_panes = "A10"
    _autofit(ws)


def _rs_sheet(ws, cfg):
    ws.sheet_view.showGridLines = False
    rs = cfg.get("rs", {})
    ws.cell(1, 1, rs.get("key_label", "Key")).font = F_BOLD
    ws.cell(1, 2, "in Cognos but not in PBI").font = F_BOLD
    r = 2
    for item in rs.get("cognos_only", []):
        ws.cell(r, 1, item.get("key", "")).font = F_BASE
        if item.get("note"):
            ws.cell(r, 2, item["note"]).font = F_BASE
        r += 1
    r += 1
    ws.cell(r, 1, "in PBI but not in Cognos").font = F_BOLD
    r += 1
    for item in rs.get("pbi_only", []):
        ws.cell(r, 1, item.get("key", "")).font = F_BASE
        if item.get("note"):
            ws.cell(r, 2, item["note"]).font = F_BASE
        r += 1
    r += 1
    for note in rs.get("research_notes", []):
        ws.cell(r, 3, note).font = F_BASE
        r += 1
    r += 2
    # column-name map: Cognos row above PBI row, like Rohit's RS!A16:I17
    cmap = rs.get("col_map")
    if cmap:
        ws.cell(r, 1, "Cognos").font = F_BOLD
        for i, h in enumerate(cmap["cognos"]):
            ws.cell(r, 2 + i, h).font = F_BASE
        r += 1
        ws.cell(r, 1, "PBI").font = F_BOLD
        for i, h in enumerate(cmap["pbi"]):
            ws.cell(r, 2 + i, h).font = F_BASE
    _autofit(ws)


def load_comparison_csv(path, n_cognos, n_flags, n_pbi, skip_header=True):
    """Split an agent comparison CSV (Cognos cols | flag cols | PBI cols,
    possibly with 1-col separators) into row tuples for _comparison_sheet."""
    rows = _read_csv(path)
    hdr = rows[0]
    data = rows[1:] if skip_header else rows
    out = []
    for row in data:
        row = row + [""] * (n_cognos + n_flags + n_pbi - len(row))
        out.append((row[:n_cognos],
                    row[n_cognos:n_cognos + n_flags],
                    row[n_cognos + n_flags:n_cognos + n_flags + n_pbi]))
    return hdr, out


def build_workbook(cfg, out_path):
    wb = Workbook()
    _notes_sheet(wb.active, cfg)
    wb.active.title = "Notes"
    for section in cfg["comparisons"]:
        title = section.get("sheet_name", "Comparison")[:31]
        _comparison_sheet(wb.create_sheet(title), cfg, section)
    _rs_sheet(wb.create_sheet("RS"), cfg)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path

# ----------------------------------------------------------------------------
# config schema:
# {
#   "report_id": 323, "report_name": "Shell and Kemper 530 Report",
#   "data_revised": "7/6/2026",
#   "notes": {
#     "filters": ["...line per Cognos filter..."],
#     "slicers": ["Owner"],
#     "links": {"powerbi": None_or_url, "cognos": None_or_url},
#     "report_location_cognos": "Michelman Reporting, Production and Shipping, ...",
#     "report_location_edw": "Michelman Inc, Power BI Reports, ...",
#     "perspective": None, "jde_live": True,
#     "comments": ["multi\nline blocks", ...]
#   },
#   "comparisons": [
#     {"sheet_name": "Comparison", "sort_note": "Sorted by ...",
#      "cognos_headers": [...], "pbi_headers": [...],
#      "rows": [(cognos_vals, flag_vals, pbi_vals), ...]}
#   ],
#   "rs": {"key_label": "PO-Vendor-Lot", "cognos_only": [{"key":..,"note":..}],
#          "pbi_only": [...], "research_notes": [...],
#          "col_map": {"cognos": [...], "pbi": [...]}}
# }
# ----------------------------------------------------------------------------
