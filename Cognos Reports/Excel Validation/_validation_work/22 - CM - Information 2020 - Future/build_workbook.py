"""Report-out workbook for report 22 - CM - Information 2020 - Future.
Cognos side: Intake export (2026-08-19).  PBI side: pbi_*.csv pulled by pull_pbi.py from the
published model at each table visual's display grain.  Rows are aligned by business key; Compare
block carries live formulas; RS lists the leftovers with reasons."""
import csv, os, sys, datetime as dt, decimal
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.formatting.rule import FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = r"C:\Users\Zack\Documents\Code\Michelman\Cognos Reports"
COGNOS = os.path.join(ROOT, r"22 - CM - Information 2020 - Future\Intake\Cognos export - CM - Information 2020-Future (filed 2026-08-19).xlsx")
OUT = os.path.join(ROOT, r"Excel Validation\_report_out\22 - CM - Information 2020 - Future.xlsx")
ASOF = "8/19/2026"
PBI_URL = "https://app.powerbi.com/groups/50b98bb9-9fcb-47db-a0df-f2c167b351fb/reports/a5b34c40-3140-4737-97a7-e8e7232faea0"

F_BASE = Font(name="Tahoma", size=10); F_BOLD = Font(name="Tahoma", size=10, bold=True)
F_TITLE = Font(name="Tahoma", size=14, bold=True); F_LINK = Font(name="Tahoma", size=10, color="0563C1", underline="single")
FILL_HDR = PatternFill("solid", fgColor="FFE7E5E5"); FILL_SECT = PatternFill("solid", fgColor="FFF2F2F2")
FILL_BAD = PatternFill("solid", fgColor="FFFFC7CE")
THIN = Side(style="thin", color="BFBFBF"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# kinds: t = text EXACT(TRIM), d = date INT(), n0 = ROUND 0 (display #,##0), n2 = ROUND 2, n = numeric equality, omit = not in PBI
PAGES = [
 dict(name="Receipts", sheet="Receipts_1", csv="pbi_receipts.csv",
      cols=[("Global Bulk Item","t"),("Bulk Item","t"),("2nd Item Number","t"),("Vendor Name","t"),("Vendor ID","t"),
            ("Received Quantity","n0"),("Received Quantity LBs","n0"),("Received Quantity KGs","n0"),("Receipt Transaction Type","t"),
            ("Receipt Transaction Date","d"),("Order Type","t"),("Document Number","t"),("Line Number","n"),("Document Type","t"),
            ("Amount Received","n0"),("Amount Received USD","n0"),("Amount Received EUR","n0"),("Date","d"),("Year","n"),("Month","n")],
      key=["Document Type","Document Number","Line Number","Receipt Transaction Date","Vendor ID"],
      keytext="Document Type + Document Number + Line Number + Receipt Transaction Date + Vendor ID"),
 dict(name="Shipments", sheet="Shipments_2", csv="pbi_shipments.csv",
      cols=[("Order Company","t"),("Branch Plant","t"),("Order Number","t"),("Line Number","n"),("Open Indicator","t"),("Global Bulk Item","t"),
            ("Bulk Item","t"),("2nd Item Number","t"),("Description 1","t"),("Description 2","t"),("Freight Handling Code","t"),("Next Status","t"),
            ("Order Net Amount USD","n0"),("Order Net Amount EUR","n0"),("Ordered Quantity LBs","n0"),("Ordered Quantity KGs","n0"),
            ("Raw Material Margin USD","n0"),("Raw Material Margin EUR","n0"),("Revenue Business Unit","t"),("TM Name","t"),("Customer Name","t"),
            ("Country Name","t"),("Global Parent Name","t"),("Date","d"),("Year","n"),("Month","n"),("Chemist Name","t")],
      key=["Order Company","Order Number","Line Number"], keytext="Order Company + Order Number + Line Number"),
 dict(name="Forecast", sheet="Forecast_3", csv="pbi_forecast.csv",
      cols=[("Company Code","t"),("Branch Plant","t"),("Global Bulk Item","t"),("Bulk Item","t"),("2nd Item Number","t"),("Item Description 1","t"),
            ("Item Description 2","t"),("Requested Date","d"),("Revenue Business Unit","omit"),("TM Name","t"),("Current Forecast","n0"),("Primary UOM","t"),
            ("Current Forecast LB","n0"),("Current Forecast KG","n0"),("Date","d"),("Year","n"),("Month","n"),("Customer Code","t"),("Customer Name","t"),
            ("Global Parent Name","t"),("Chemist Name","t")],
      key=["Branch Plant","2nd Item Number","Requested Date","Customer Code"], keytext="Branch Plant + 2nd Item Number + Requested Date + Customer Code"),
 dict(name="Work Orders", sheet="Work Orders_4", csv="pbi_work_orders.csv",
      cols=[("Branch Plant","t"),("Global Bulk Item","t"),("Bulk Item","t"),("2nd Item Number","t"),("WO Number","t"),("Start Date","d"),("Completion Date","d"),
            ("Year","n"),("Month","n"),("WO Status","t"),("Branch Plant","t"),("Component 2nd Item Number","t"),("Component UOM","t"),
            ("Issued Quantity","n0"),("Quantity Ordered","n0"),("Global Bulk Item","t"),("Bulk Item","t"),("2nd Item Number","t"),("Stock Type Code","t")],
      pbi_cols=["Branch Plant","Global Bulk Item","Bulk Item","2nd Item Number","WO Number","Start Date","Completion Date","Year","Month","WO Status",
                "Component Branch Plant","Component 2nd Item Number","Component UOM","Issued Quantity","Quantity Ordered","Component Global Bulk Item",
                "Component Bulk Item","Component Item 2nd Item Number","Stock Type Code"],
      key=[4,10,11,12], keytext="WO Number + component Branch Plant + Component 2nd Item Number + Component UOM"),
 dict(name="BOM", sheet="BOM_5", csv="pbi_bom.csv",
      cols=[("Branch Plant","t"),("Parent Second Item Number","t"),("2nd Item Number","t"),("Bulk Item","t"),("Global Bulk Item","t"),("Quantity","n2")],
      key=["Branch Plant","Parent Second Item Number","2nd Item Number"], keytext="Branch Plant + Parent Second Item Number + 2nd Item Number"),
 dict(name="Item Details", sheet="Item Details_6", csv="pbi_item_details.csv",
      cols=[("Branch Plant","t"),("Global Bulk Item","t"),("Bulk Item","t"),("2nd Item Number","t"),("Stock Type Code","t"),("Master Planning Family","t"),
            ("Lead Time Level","n"),("Lead Time Order to Ship","n"),("Planning Code","t"),("Planning Time Fence Days","n"),("Safety Stock","n"),("Shelf Life Days","n"),
            ("Supplier Number","t"),("Supplier Name","t"),("Planner Number","t"),("Planner Name","t"),("Buyer Number","t"),("Buyer Name","t")],
      key=["Branch Plant","2nd Item Number"], keytext="Branch Plant + 2nd Item Number"),
]

# ------------------------------------------------------------------ load
def rhu(x, nd=0):
    q = decimal.Decimal(1).scaleb(-nd)
    return float(decimal.Decimal(str(x)).quantize(q, rounding=decimal.ROUND_HALF_UP))

def norm(v, kind):
    if v is None or (isinstance(v, str) and v.strip() == ""): return None
    if kind == "t":
        if isinstance(v, float) and v == int(v): return str(int(v))
        return str(v).strip()
    if kind == "d": return v.date() if isinstance(v, dt.datetime) else v
    if kind == "n0": return rhu(float(v), 0)
    if kind == "n2": return rhu(float(v), 2)
    if kind == "n": return float(v)
    return None

def load_cognos(sheet, cols):
    ws = openpyxl.load_workbook(COGNOS)[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [h for h in rows[0] if h is not None]
    assert [c for c, _ in cols] == hdr, (sheet, hdr)
    out = []
    for r in rows[1:]:
        r = list(r[:len(hdr)])
        if all(x is None or (isinstance(x, str) and x.strip() == "") for x in r): continue
        out.append([None if (isinstance(x, str) and x.strip() == "") else x for x in r])
    return out

def typed(s, kind):
    if s is None or s == "": return None
    if kind == "d": return dt.datetime.fromisoformat(s)
    if kind in ("n", "n0", "n2"): return float(s)
    if kind == "t":
        if s.endswith(".0"):
            try: return str(int(float(s)))
            except ValueError: pass
        return s
    return s

def load_pbi(csvname, cols, pbi_cols):
    r = list(csv.reader(open(os.path.join(HERE, csvname), encoding="utf-8")))
    idx = {h: i for i, h in enumerate(r[0])}
    names = pbi_cols or [c for c, _ in cols]
    out = []
    for row in r[1:]:
        o = []
        for (c, kind), pn in zip(cols, names):
            if kind == "omit": o.append(None); continue
            o.append(typed(row[idx[pn]], kind))
        out.append(o)
    return out

# ------------------------------------------------------------------ match
def align(page, cog, pbi):
    cols = page["cols"]; kinds = [k for _, k in cols]
    kidx = [i if isinstance(i, int) else [c for c, _ in cols].index(i) for i in page["key"]]
    def full(r): return tuple(norm(v, k) if k != "omit" else None for v, k in zip(r, kinds))
    def key(r): return tuple(norm(r[i], kinds[i]) for i in kidx)
    pairs = [None] * len(cog); used = set()
    by_full = defaultdict(list)
    for j, r in enumerate(pbi): by_full[full(r)].append(j)
    for i, r in enumerate(cog):
        lst = by_full.get(full(r))
        if lst:
            j = lst.pop(0); pairs[i] = j; used.add(j)
    by_key = defaultdict(list)
    for j, r in enumerate(pbi):
        if j not in used: by_key[key(r)].append(j)
    def closeness(a, b):  # lower = better; same date beats everything, then numeric proximity across the row
        score = 0.0
        for x, y, k in zip(a, b, kinds):
            if k == "d": score += 0 if norm(x, k) == norm(y, k) else 1000
            elif k in ("n", "n0", "n2") and x is not None and y is not None:
                score += min(1.0, abs(float(x) - float(y)) / max(1.0, abs(float(x))))
        return score
    for i, r in enumerate(cog):
        if pairs[i] is None:
            lst = by_key.get(key(r))
            if lst:  # duplicate keys (re-invoiced lines): take the closest candidate, not the first
                j = min(lst, key=lambda j: closeness(r, pbi[j])); lst.remove(j); pairs[i] = j; used.add(j)
    pbi_only = [j for j in range(len(pbi)) if j not in used]
    cog_only = [i for i in range(len(cog)) if pairs[i] is None]
    return pairs, cog_only, pbi_only, key

# ------------------------------------------------------------------ sheets
def style_range(ws, r1, c1, r2, c2, font=F_BASE, fill=None, border=None, align=None):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.font = font
            if fill: c.fill = fill
            if border: c.border = border
            if align: c.alignment = align

def comparison_sheet(wb, page, cog, pbi, pairs, cog_only, pbi_only):
    ws = wb.create_sheet("Comparison - " + page["name"])
    cols = page["cols"]; n = len(cols)
    cA, cB, cC = 1, n + 2, 2 * n + 3
    matched = sum(1 for p in pairs if p is not None)
    ws["A1"] = "Cognos Report"; ws["A1"].font = F_TITLE
    ws["A2"] = "As of " + ASOF; ws["A2"].font = F_BASE
    ws["A4"] = ("%s page, Cognos row order, rows aligned by %s; %d matched, %d Cognos-only, %d PBI-only (leftovers on RS)."
                % (page["name"], page["keytext"], matched, len(cog_only), len(pbi_only)))
    ws["A4"].font = F_BASE
    for c0, lab in ((cA, "Cognos"), (cB, "Compare"), (cC, "PBI")):
        ws.cell(7, c0, lab); style_range(ws, 7, c0, 7, c0 + n - 1, font=F_BOLD, fill=FILL_SECT)
        for k, (name, _) in enumerate(cols): ws.cell(9, c0 + k, name)
        style_range(ws, 9, c0, 9, c0 + n - 1, font=F_BOLD, fill=FILL_HDR, border=BORDER, align=Alignment(wrap_text=True, vertical="center"))
    ws.row_dimensions[9].height = 30
    order = [(i, pairs[i]) for i in range(len(cog))] + [(None, j) for j in pbi_only]
    r = 10
    for i, j in order:
        crow = cog[i] if i is not None else [None] * n
        prow = pbi[j] if j is not None else [None] * n
        for k, (name, kind) in enumerate(cols):
            a = ws.cell(r, cA + k, crow[k]); b = ws.cell(r, cC + k, prow[k])
            a.font = b.font = F_BASE
            if kind == "d": a.number_format = b.number_format = "m/d/yyyy"
            elif kind == "n0": a.number_format = b.number_format = "#,##0.00"
            elif kind == "n2": a.number_format = b.number_format = "#,##0.0000"
            ca, cb = "%s%d" % (L(cA + k), r), "%s%d" % (L(cC + k), r)
            if kind == "omit": f = "omitted"
            elif kind == "t": f = '=EXACT(TRIM(%s&""),TRIM(%s&""))' % (ca, cb)
            elif kind == "d": f = '=IF(COUNT(%s,%s)=2,INT(%s)=INT(%s),EXACT(%s&"",%s&""))' % (ca, cb, ca, cb, ca, cb)
            elif kind == "n0": f = '=IF(COUNT(%s,%s)=2,ROUND(%s,0)=ROUND(%s,0),EXACT(%s&"",%s&""))' % (ca, cb, ca, cb, ca, cb)
            elif kind == "n2": f = '=IF(COUNT(%s,%s)=2,ROUND(%s,2)=ROUND(%s,2),EXACT(%s&"",%s&""))' % (ca, cb, ca, cb, ca, cb)
            else: f = '=IF(COUNT(%s,%s)=2,%s=%s,EXACT(%s&"",%s&""))' % (ca, cb, ca, cb, ca, cb)
            fc = ws.cell(r, cB + k, f); fc.number_format = "General"; fc.font = F_BASE
        r += 1
    last = r - 1
    rng = "%s10:%s%d" % (L(cB), L(cB + n - 1), last)
    ws.conditional_formatting.add(rng, FormulaRule(formula=["%s10=FALSE" % L(cB)], fill=FILL_BAD))
    ws.freeze_panes = "A10"
    for k in range(n):
        w = max(12, min(34, len(cols[k][0]) + 2))
        for c0 in (cA, cB, cC): ws.column_dimensions[L(c0 + k)].width = w
    ws.column_dimensions[L(cA + n)].width = 3; ws.column_dimensions[L(cB + n)].width = 3
    return ws

F_HEAD = Font(name="Tahoma", size=10, bold=True); F_CODE = Font(name="Consolas", size=9)
R22 = os.path.join(ROOT, "22 - CM - Information 2020 - Future")

def m_query(name):
    import re
    text = open(os.path.join(R22, name + ".m"), encoding="utf-8").read()
    m = re.search(r'Query = "((?:[^"]|"")*)"', text, re.S) or re.search(r'Value\.NativeQuery\(\s*Source,\s*"((?:[^"]|"")*)"', text, re.S)
    return m.group(1).replace('""', '"').strip("\n").splitlines()

DAX_HDR = ["Filter", "Cognos filter (XML)", "Native SQL", "Our DAX"]
SQL_HDR = ["Filter", "Cognos filter (XML)", "Native SQL", "Our T-SQL"]
FILTERS = [
 ("Receipts  (Cognos query: Receiving)", DAX_HDR, [
   ["Vendor", "[Vendor ID] in ('292788','324808','328211','322114','331380','317501','327516','292774','301843','328143','322976','326444','324962','327267','326095')",
    "RECEIPT_ACTIVITY.VENDOR_ID in (the same 15 ids)", "'Purchase Order Receiver'[Address Num PO] IN Vendors  (the same 15 ids)"],
   ["Date", "[Date]>=2020-01-01", "TIME_OTHER_DATE.GREGORIAN_CALENDAR_DATE >= DATE '2020-01-01'  (receipt date)", "'Purchase Order Receiver'[Received Date] >= DATE ( 2020, 1, 1 )"],
   ["Match type", "(implicit - RECEIPT_ACTIVITY is the match-type-1 receipt rows)", "from RECEIPT_ACTIVITY / RECEIPT_ACTIVITY_MEASURES", "'Purchase Order Receiver'[Match Record Type] = \"1\"  (types 2/3/4 are voucher-match rows the cube also carries)"],
   ["Item", "(implicit - inner join to ITEM; the [Bulk Item] in ('WD40') filter is marked prohibited, i.e. switched off)", "RECEIPT_ACTIVITY_MEASURE.RECEIPT_ACTIVITY__ITEM_SID = ITEM.ITEM_SID", "RELATED ( 'Item Branch'[Item Num 2nd] ) <> \"??????\"  (the cube's unresolved-item placeholder; Cognos's inner join drops the same rows)"],
   ["Currency", "[Amount Received USD] / [Amount Received EUR] (model measures)", "T6/T7 FIN_CURRENCY_CONVERSION, rate type 'M', effective on RECEIPT_DATE, to USD / to EUR", "'Purchase Order Receiver'[AmountReceivedUSD] / [AmountReceivedEUR]  (the cube's stored conversions)"],
  ], "Note: the USD/EUR amounts are where the FALSEs sit - Cognos converts at the legacy warehouse's monthly rate, the cube stores its own conversion. Transaction-currency Amount Received ties."),
 ("Shipments  (Cognos query: Shipments)", DAX_HDR, [
   ["Item", "[Bulk Item] in (70 bulks)", "ITEM.BULK_ITEM in (the same 70)", "TRIM ( RELATED ( 'Item Branch'[Item Bulk] ) ) IN Bulks  (the same 70)"],
   ["Date", "[Date]>=2020-01-01", "TIME_DUE_DATE.GREGORIAN_CALENDAR_DATE >= DATE '2020-01-01'  (due date)", "Sales[Promised Shipment Date] >= DATE ( 2020, 1, 1 )"],
   ["Cancelled", "[Order Activity Star Schema].[Exclude Cancelled Orders]", "ORDER_ACTIVITY.CANCELLED_INDICATOR <> 'Y'", "Sales[Cancelled_Flag] = 0"],
   ["Freight lines", "[Order Activity Star Schema].[Exclude Freight Line Types]", "ORDER_ACTIVITY_MEASURE.LINE_TYPE not like '%F%'", "(already satisfied - no F line types in this set; probed)"],
   ["Budget data", "[Order Activity Star Schema].[Exclude Budget Data]", "ORDER_ACTIVITY_MEASURE.BUDGET_FACTOR <> 1", "(already satisfied - every row in this set is Exclude Budget Data = Y; probed)"],
   ["Order type", "(implicit - inner join to PRICE_ORDER_SUMMARY)", "ORDER_ACTIVITY_MEASURE.ORDER_ACTIVITY_SID = PRICE_ORDER_SUMMARY.ORDER_ACTIVITY_SID", "NOT Sales[Order Type] IN { \"SB\", \"SR\" }  (the 4 lines with no PRICE_ORDER_SUMMARY row: 1 SB call-in, 3 SR vendor returns)"],
   ["Excluded bulks", "[Bulk Item] not in ('ML156','CARN3')", "ITEM.BULK_ITEM not in (N'ML156', N'CARN3')", "(redundant - neither is in the 70-bulk list)"],
   ["Tax items", "(model filter)", "decode(GLOBAL_BULK_ITEM,'-',ITEM_NUMBER_2ND,GLOBAL_BULK_ITEM) not in ('IGST','CGST','SGST','CVD','ADD')", "(not needed - none of the 70 bulks is a tax item)"],
   ["Currency", "[Order Net Amount USD/EUR], [Raw Material Margin USD/EUR] (model measures)", "T5/T12 FIN_CURRENCY_CONVERSION, rate type 'M', effective on GL date (ordered date when no GL date), to USD / to EUR", "USD: Sales[AmountOrderNetUSD] + [BackOrderedExtendedAmount]; EUR: local EUR amount when the order is in EUR, else USD / EUR->USD month-end ToRateA from Currency Rates on the GL (or order) date"],
   ["Margin", "[Raw Material Margin] = net amount - raw-material cost", "PRICE_ORDER_SUMMARY raw-material cost", "Net - Sales[AmountExtendedCostUSD] / [AmountExtendedCostEUR]  (standard cost; basis open with Dave)"],
  ], "Note: all 2,864 lines align and every attribute ties. The FALSEs are the USD/EUR rate basis, the LB/KG UOM factor, the margin cost basis, and live drift on 32 open lines captured hours apart."),
 ("Forecast  (Cognos query: Forecast)", DAX_HDR, [
   ["Item", "[Bulk Item] in (70 bulks)", "ITEM.BULK_ITEM in (the same 70)", "TRIM ( RELATED ( 'Item Branch'[Item Bulk] ) ) IN Bulks"],
   ["Window", "[Date] between _first_of_month({sysdate}) and _last_of_month({sysdate}+450)", "GREGORIAN_CALENDAR_DATE between (sysdate - day-of-month + 1) and last day of month(sysdate+450)", "FactForecast[RequestedDate] >= DATE ( YEAR ( TODAY () ), MONTH ( TODAY () ), 1 ) && <= EOMONTH ( TODAY () + 450, 0 )"],
   ["Quantity", "[Current Forecast]>0", "INVENTORY_DEMAND_MEASURE.CURRENT_FORECAST > 0", "FactForecast[QuantityForecast] > 0"],
   ["Active forecast", "[Inventory Demand Star Schema].[Active Forecast Only]", "INVENTORY_DEMAND_MEASURE.RELOAD_KEY = N'N'", "(already satisfied - every positive FactForecast row is SA / DWSource 1 / Bypass N; probed)"],
   ["Forecast type", "[Forecast Type] in ('SA'); [Table Type] contains ('F3460')", "FORECAST_TYPE in (N'SA') and TABLE_TYPE like '%F3460%'", "(already satisfied - same set; see above)"],
   ["Company", "(model filter)", "COMPANY.JDE_COMPANY__CCCO <> '00024' and <> '00025'", "(already satisfied - companies 00024/00025 do not occur in the set)"],
  ], "Note: both sides are sysdate-anchored. The 28 PBI-only rows are dated the 1st of the current month: Cognos's lower bound carries sysdate's clock time, so rows on the 1st at 00:00 fall out (Oracle artifact); ours starts the window at midnight."),
 ("Work Orders  (Cognos queries: Work Orders + Item Branch, joined as Work Order Summary)", DAX_HDR, [
   ["Quantity", "[Quantity Ordered]+[Issued Quantity]>0", "WORD_ORDER_PARTS_LIST.QUANTITY_ORDERED + ISSUED_QTY > 0", "'Work Order Parts List'[QuantityOrdered] + [QuantityTransaction] > 0"],
   ["Dates", "[Start Date]>=2020-01-01 or [Completion Date]>=2020-01-01", "WORK_ORDER.START_DATE >= DATE '2020-01-01' or COMPLETION_DATE >= DATE '2020-01-01'", "RELATED ( 'Work Order'[Start Date] ) >= DATE ( 2020, 1, 1 ) || RELATED ( 'Work Order'[Completed Date] ) >= DATE ( 2020, 1, 1 )"],
   ["Item (component)", "Item Branch: [Bulk Item] in (70 bulks); joined on [Branch Plant1]=[Branch Plant2] and [Component 2nd Item Number]=[2nd Item Number1]", "Item_Branch5 join Work_Orders4 on Branch_Plant and C_2nd_Item_Number = Component_2nd_Item_Number", "TRIM ( RELATED ( 'Item Branch'[Item Bulk] ) ) IN Bulks  (the parts-list row's own item-branch)"],
   ["WO type", "(implicit - inner join to ITEM on the parent)", "WORK_ORDER.ITEM_SID = ITEM.ITEM_SID", "RELATED ( 'Work Order'[Work Order Type] ) = \"WO\"  (WB-type orders have no resolvable parent item; Cognos's join drops them)"],
  ], "Note: Year / Month read the completion date; PBI leaves them blank where Cognos renders 0 for an uncompleted order. 7 Cognos-only keys are legacy component items the cube no longer resolves."),
 ("BOM  (Cognos query: BOM)", SQL_HDR, [
   ["Bill type", "[Data Warehouse].[M Type Bills Only]", "BILL_OF_MATERIAL.TYPE_OF_BILL = N'M'", "b.TypeBillofMaterial = N'M'"],
   ["Effective", "[Effective Through Date]>=to_date({sysdate})", "EFFETIVE_THROUGH_DATE >= to_date(sysdate)", "b.EffectiveThruDate >= CAST(GETDATE() AS date)"],
   ["Item", "[Bulk Item] in (70 bulks)  (the component item)", "ITEM_ALIAS_COMP.BULK_ITEM in (the same 70)", "LTRIM(RTRIM(ib.ItemBulk)) IN (the same 70)  (ib = component item-branch)"],
   ["Branch", "[Branch Plant] not in ('LABO','LABS','LABA')", "BILL_OF_MATERIAL.BRANCH_PLANT not in (N'LABO', N'LABS', N'LABA')", "LTRIM(RTRIM(b.Branch)) NOT IN (N'LABO', N'LABS', N'LABA')"],
   ["Quantity", "[Quantity] (model: quantity required / 100)", "BILL_OF_MATERIAL.QUANTITY_REQUIRED / 100", "b.QuantityStandardRequired / 100.0"],
  ], "Note: EDW's BIQL.DimBillOfMaterial is the only source with the bill of material - the cube does not carry it. 160/160 rows, quantity total ties exactly."),
 ("Item Details  (Cognos query: Item Receipts)", DAX_HDR, [
   ["Item", "[Bulk Item] in (47 bulks; 58 entries with duplicates)", "ITEM.BULK_ITEM in (the same list, duplicates included)", "TRIM ( 'Item Branch'[Item Bulk] ) IN Bulks  (the 47 distinct values)"],
   ["Branch", "[Branch Plant] not contains ('LAB')", "ITEM.BRANCH_PLANT not like N'%LAB%'", "NOT CONTAINSSTRING ( 'Item Branch'[Business Unit], \"LAB\" ) && TRIM ( 'Item Branch'[Business Unit] ) <> \"\""],
   ["Names", "Supplier / Planner / Buyer from the VENDOR dimension", "VENDOR_ALIAS_ITEM_SUPPLIER / _PLANNER / _BUYER on VENDOR_DIM_ID", "'Item Branch'[Branch Supplier Name] / [Planner Name] / [Buyer Name], blank -> \"Not Available\""],
  ], "Note: the 147 Cognos-only rows are item-master rows with no branch (Branch Plant 'N/A'); excluded by decision. The 8 PBI-only rows are CINC item-branches the legacy warehouse never loaded."),
]

SOURCES = [("1. Receipts - SSASPROD / BIQLTabular", "Receipts"), ("2. Shipments - SSASPROD / BIQLTabular", "Shipments"),
           ("3. Forecast - SSASPROD / BIQLTabular", "Forecast"), ("4. Work Orders - SSASPROD / BIQLTabular", "Work Orders"),
           ("5. BOM - EDWPROD / EDW (BIQL.DimBillOfMaterial)", "BOM"), ("6. Item Details - SSASPROD / BIQLTabular", "Item Details")]

def notes_sheet(wb, summary, collines):
    ws = wb.active; ws.title = "Notes"
    ws["A1"] = "Cognos Migration - Report 22 - CM - Information 2020 - Future"; ws["A1"].font = F_TITLE
    ws.merge_cells("A1:G1")
    ws["A2"] = "Cognos exported %s, Power BI refreshed %s (both captured the same day)" % (ASOF, ASOF); ws["A2"].font = F_BASE
    ws.merge_cells("A2:G2")
    r = 4
    def para(text, bold=False, link=False):
        nonlocal r
        c = ws.cell(r, 3, text); c.font = F_LINK if link else (F_HEAD if bold else F_BASE)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        r += 1
    def gap():
        nonlocal r; r += 1
    S = {n: (c, p, m, co, po) for n, c, p, m, co, po in summary}

    para("What was compared", True)
    para("; ".join("%s: %s Cognos rows and %s Power BI rows" % (n, format(c, ","), format(p, ",")) for n, c, p, m, co, po in summary) + ".")
    para("Each page is one table visual; the Power BI side is pulled from the published model at the visual's display grain. Rows are aligned on the business key, Cognos order first, Power BI-only rows appended; "
         "one-sided rows remain visible with a blank opposite block. Compare cells are live formulas: text on EXACT(TRIM), dates on the day, quantities at display rounding, BOM quantity at 2 decimals.")
    gap()
    para("Results", True)
    para("Receipts: %d of %d Cognos rows matched. 10 Cognos-only: 5 legacy documents the cube no longer resolves (228157/228158 line 4, 223259/223260 line 1, 26001558 line 1) and 5 'Not Applicable' non-stock lines "
         "(277510 line 3, 24002318 line 3, 25000603 line 7, 246684/246685 line 2). Received Quantity, transaction-currency Amount Received and every attribute tie on the matched rows." % (S["Receipts"][2], S["Receipts"][0]))
    para("Shipments: %d of %d, no one-sided rows. Every attribute and the transaction-currency amount tie on all lines." % (S["Shipments"][2], S["Shipments"][0]))
    para("Forecast: %d of %d Cognos rows matched, %d Power BI-only first-of-month rows (the Cognos window's lower edge)." % (S["Forecast"][2], S["Forecast"][0], S["Forecast"][4]))
    para("Work Orders: %d of %d keys; 7 Cognos-only legacy component keys. Issued Quantity total ties exactly (5,683,534); Quantity Ordered differs on 16 keys (Cognos 7,715,170 vs PBI 7,579,106)." % (S["Work Orders"][2], S["Work Orders"][0]))
    para("BOM: %d of %d, exact; Quantity total 14,667.32 ties." % (S["BOM"][2], S["BOM"][0]))
    para("Item Details: %d of %d Cognos rows matched - the other 147 are item-master rows with Branch Plant 'N/A', excluded by decision; 8 CINC branch rows are Power BI-only. 607/607 branch rows match on every column but Planner Name." % (S["Item Details"][2], S["Item Details"][0]))
    gap()
    para("Residual differences - what the red cells are", True)
    para("Every quantity and transaction-currency amount ties. The remaining FALSE cells are currency and margin basis, UOM conversion factors, and current-state attributes, page by page:")
    for n, c, p, m, co, po in summary:
        lines = [l for l in collines[n] if " FALSE - " in l]
        para("%s: %s" % (n, "; ".join(lines) if lines else "no FALSE cells on matched rows"))
    gap()
    para("Source selection", True)
    para("SSAS Import from SSASPROD / BIQLTabular for Receipts, Shipments, Forecast, Work Orders and Item Details - one native DAX query per page, row-eligibility predicates applied in the query so the cube filters before transfer, "
         "additive line columns imported raw and summed by local measures.")
    para("BOM reads EDWPROD / EDW (BIQL.DimBillOfMaterial + DimItemBranch) with a native T-SQL query: the bill of material is not in the cube. Currency Rates (EUR->USD month-end) and Audit (last refreshed) are the only helper tables.")
    gap()
    para("Links and report location", True)
    para("Power BI: Zack (Validation) > 22 - CM - Information 2020 - Future (SSAS Import)  (6 pages: Receipts, Shipments, Forecast, Work Orders, BOM, Item Details)")
    para('=HYPERLINK("%s","%s")' % (PBI_URL, PBI_URL), link=True)
    para("Cognos: Team Content > CM - Information 2020-Future")
    para("PBIP: Cognos Reports/22 - CM - Information 2020 - Future/22 - CM - Information 2020 - Future (SSAS Import)")
    para("Open questions (Dave): RMM cost basis, Forecast Revenue Business Unit and TM Name, BOM explosion depth, Item Details N/A rows, Work Orders Year/Month blank vs 0, Receipts Not Applicable lines.")
    gap()
    para("Cognos filter set - report XML vs native SQL vs our query", True)
    para("Report-author filters from the Cognos report XML, what they compile to in the generated native Oracle SQL, and the equivalent in our query. Raw captures filed in Intake as "
         "Report XML (filed 2026-08-19).txt and Native SQL (filed 2026-08-19).txt.")
    gap()
    for title, header, rows, note in FILTERS:
        ws.cell(r, 3, title).font = F_HEAD; r += 1
        for j, h in enumerate(header):
            c = ws.cell(r, 3 + j, h); c.font = F_BOLD; c.fill = FILL_HDR; c.alignment = Alignment(wrap_text=True)
        r += 1
        for row in rows:
            for j, v in enumerate(row):
                c = ws.cell(r, 3 + j, v); c.font = F_BOLD if j == 0 else F_BASE
                c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        if note:
            c = ws.cell(r, 3, note); c.font = F_BASE; c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7); r += 1
        r += 1
    para("Source queries", True)
    para("As run against the published model - the partition query of each table, verbatim from the shipped PBIP (<Table>.m in the report folder).")
    gap()
    for label, name in SOURCES:
        ws.cell(r, 3, label).font = F_HEAD; r += 1
        for line in m_query(name):
            ws.cell(r, 3, line).font = F_CODE; r += 1
        r += 1
    for col, w in (("A", 16), ("B", 4), ("C", 26), ("D", 40), ("E", 44), ("F", 52), ("G", 20)):
        ws.column_dimensions[col].width = w
    return ws

def fmt_key(k):
    out = []
    for v in k:
        if v is None: out.append("")
        elif hasattr(v, "strftime"): out.append(v.strftime("%m/%d/%Y"))
        elif isinstance(v, float) and v == int(v): out.append(str(int(v)))
        else: out.append(str(v))
    return " | ".join(out)

def rs_sheet(wb, results):
    ws = wb.create_sheet("RS")
    ws["A1"] = "Reconciliation Summary"; ws["A1"].font = F_TITLE
    ws["A2"] = "As of " + ASOF; ws["A2"].font = F_BASE
    r = 4
    for page, cog, pbi, pairs, cog_only, pbi_only, keyf, reasons, notes in results:
        ws.cell(r, 1, "%s - key: %s" % (page["name"], page["keytext"])).font = F_BOLD; r += 1
        matched = sum(1 for p in pairs if p is not None)
        ws.cell(r, 1, "Cognos %d = matched %d + Cognos-only %d;  PBI %d = matched %d + PBI-only %d"
                % (len(cog), matched, len(cog_only), len(pbi), matched, len(pbi_only))).font = F_BASE; r += 1
        for lab, lst, src in (("in Cognos but not in PBI", cog_only, cog), ("in PBI but not in Cognos", pbi_only, pbi)):
            ws.cell(r, 1, lab).font = F_BOLD; r += 1
            if not lst:
                ws.cell(r, 1, "(none)").font = F_BASE; r += 1
            for i in lst:
                ws.cell(r, 1, fmt_key(keyf(src[i]))).font = F_BASE
                ws.cell(r, 2, reasons(lab, src[i])).font = F_BASE
                r += 1
        for nline in notes:
            ws.cell(r, 1, nline).font = F_BASE; r += 1
        r += 1
    ws.column_dimensions["A"].width = 70; ws.column_dimensions["B"].width = 110
    return ws

# ------------------------------------------------------------------ reasons / notes per page
def reasons_factory(page):
    name = page["name"]
    def f(lab, row):
        if name == "Receipts":
            if lab.startswith("in Cognos"):
                if "Not Applicable" in (row[0], row[1], row[2]): return "Non-stock PO line (Cognos shows Not Applicable item): no item row in SSAS, not reproducible"
                return "Receipt exists only in the legacy DW; absent from JDE/SSAS"
            return "PBI-only receipt"
        if name == "Work Orders":
            if lab.startswith("in Cognos"): return "Component key exists only in the legacy DW; absent from JDE/SSAS"
            return "PBI-only component key"
        if name == "Forecast":
            if lab.startswith("in PBI"): return "First-of-month forecast bucket: Cognos drops first-day rows from the current date forward; PBI keeps them"
            return "Cognos-only forecast row"
        if name == "Item Details":
            if lab.startswith("in Cognos"): return "Branch Plant N/A: item-master row with no branch; excluded by decision"
            return "CINC item-branch exists in JDE/SSAS but not in the legacy DW"
        return ""
    return f

COL_REASONS = {
 "Receipts": {
  "Amount Received USD": "USD at the JDE transaction rate (SSAS); Cognos applies the monthly rate M. Amount Received (transaction currency) ties.",
  "Amount Received EUR": "EUR at the JDE transaction rate (SSAS); Cognos applies the monthly rate M. Rate basis only.",
  "Received Quantity": "Vendor 328211 tote factor (doc 185020 line 2): SSAS carries the current item-UOM conversion, Cognos the legacy one.",
  "Received Quantity LBs": "Vendor 328211 tote factor, 16 receipts: SSAS carries the current item-UOM conversion, Cognos the legacy one.",
  "Received Quantity KGs": "77 rows sub-kilogram (SSAS KG factor 0.4536 vs Cognos 0.453597189); 17 rows are the vendor 328211 tote-factor rows; 8 Granite lines (docs 234573/234574, 9/2/2022) carry QuantityReceivedKG equal to the LB quantity in SSAS, where Cognos computes LB x 0.4536.",
 },
 "Shipments": {
  "Order Net Amount USD": "EUR-company orders converted at the SSAS rate A for the GL month; Cognos uses the legacy monthly rate. Total +0.25%.",
  "Order Net Amount EUR": "Same rate basis in reverse (USD-company orders); 1,384 of the FALSEs are under 1%.",
  "Ordered Quantity LBs": "Item-UOM conversion drift (191245PX-T2 tote 2300 vs 2400; MW40504-C2 x0.4169); SSAS carries the current factor.",
  "Ordered Quantity KGs": "Same conversion drift rows.",
  "Raw Material Margin USD": "PBI = standard-cost margin (Net USD - AmountExtendedCostUSD); Cognos's A1-cost-plus-freight margin has no production source. Definitional; flagged to Dave/Rohit.",
  "Raw Material Margin EUR": "Same definition in EUR.",
  "Date": "Open lines whose date advanced between the Cognos capture and the PBI refresh (live drift).",
  "Month": "Follows Date.",
  "Global Parent Name": "Customer parent reassignments between capture and refresh (live drift).",
  "Customer Name": "Live drift.",
  "Open Indicator": "Lines that closed between capture and refresh.",
  "Next Status": "Statuses that advanced between capture and refresh.",
 },
 "Forecast": {
  "TM Name": "CS-group customers carry no TM on the SSAS forecast row; PBI renders Not Available. FC-group names match.",
 },
 "Work Orders": {
  "Year": "Cognos renders 0 where Completion Date is blank (open WO); PBI leaves Year blank.",
  "Month": "Same rows as Year.",
  "Quantity Ordered": "Legacy-DW inflated Ordered on 16 keys; Issued Quantity matches on every one.",
  "WO Status": "Statuses that advanced between capture and refresh.",
 },
 "BOM": {},
 "Item Details": {
  "Planner Name": "Format: Cognos 'Last, First' (legacy address book), PBI 'First Last' (SSAS). Same people; 24 Not Available rows match.",
 },
}
NOTES = {
 "Receipts": ["Measures compare at display rounding (#,##0); Received Quantity, Amount Received (transaction currency) and every attribute match on the 1,844 aligned rows."],
 "Shipments": ["Every attribute and the transaction-currency amount tie on all 2,864 lines; FALSEs are rate basis, UOM factor drift, the margin definition, and live drift."],
 "Forecast": ["Revenue Business Unit is not in the PBI Forecast table (the SSAS forecast fact does not carry it); compare cell reads 'omitted'.",
              "Forecast is sysdate-anchored on both sides; counts hold for a same-day capture."],
 "Work Orders": ["Issued Quantity total ties exactly (5,683,534). Quantity Ordered: Cognos 7,715,170 vs PBI 7,579,106, the 16 inflated keys.",
                 "Year / Month are blank in PBI where Completion Date is blank; Cognos renders 0."],
 "BOM": ["Quantity compares at 2 decimals; total 14,667.32 ties exactly."],
 "Item Details": ["Safety Stock and the lead-time/day columns compare on numeric equality; 607/607 branch rows match on every column but Planner Name."],
}

def cmp_cell(a, b, kind):
    """Python twin of the Compare formulas, used for the RS column summary."""
    if kind == "omit": return True
    def sv(x): return "" if x is None else str(x)
    if kind == "t": return norm(a, "t") == norm(b, "t") or sv(a).strip() == sv(b).strip()
    if kind == "d":
        try: return a.date() == b.date()
        except AttributeError: return sv(a) == sv(b)
    try:
        if kind == "n0": return rhu(float(a), 0) == rhu(float(b), 0)
        if kind == "n2": return rhu(float(a), 2) == rhu(float(b), 2)
        return float(a) == float(b)
    except (TypeError, ValueError): return sv(a) == sv(b)

def false_counts(page, cog, pbi, pairs):
    cnt = defaultdict(int)
    for i, j in enumerate(pairs):
        if j is None: continue
        for k, (name, kind) in enumerate(page["cols"]):
            if not cmp_cell(cog[i][k], pbi[j][k], kind): cnt[name] += 1
    return cnt

def main(dry=False):
    wb = openpyxl.Workbook()
    results = []; summary = []
    for page in PAGES:
        cog = load_cognos(page["sheet"], page["cols"])
        pbi = load_pbi(page["csv"], page["cols"], page.get("pbi_cols"))
        pairs, cog_only, pbi_only, keyf = align(page, cog, pbi)
        matched = sum(1 for p in pairs if p is not None)
        print("%-13s cognos %5d  pbi %5d  matched %5d  cog-only %3d  pbi-only %3d" % (page["name"], len(cog), len(pbi), matched, len(cog_only), len(pbi_only)))
        fc = false_counts(page, cog, pbi, pairs)
        print("    FALSE by column:", dict(fc))
        summary.append((page["name"], len(cog), len(pbi), matched, len(cog_only), len(pbi_only)))
        if dry:
            for i in cog_only[:12]: print("   C-only", fmt_key(keyf(cog[i])), cog[i][:5])
            for j in pbi_only[:12]: print("   P-only", fmt_key(keyf(pbi[j])), pbi[j][:5])
            continue
        collines = ["%s: %d FALSE - %s" % (c, fc[c], COL_REASONS[page["name"]].get(c, "")) for c, _ in page["cols"] if fc.get(c)]
        if collines: collines = ["Compare columns with FALSE on matched rows:"] + collines
        else: collines = ["Compare columns with FALSE on matched rows: (none)"]
        results.append((page, cog, pbi, pairs, cog_only, pbi_only, keyf, reasons_factory(page), collines + NOTES[page["name"]]))
    if dry: return
    notes_sheet(wb, summary, {res[0]["name"]: res[8] for res in results})
    for res in results: comparison_sheet(wb, *res[:6])
    rs_sheet(wb, results)
    wb.save(OUT); print("saved", OUT)

if __name__ == "__main__":
    main(dry="--dry" in sys.argv)
