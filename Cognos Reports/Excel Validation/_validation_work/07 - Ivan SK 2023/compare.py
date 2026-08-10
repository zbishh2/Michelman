import openpyxl, csv, datetime, os
from collections import Counter, defaultdict

OUT = r'C:/Users/Zack/Documents/Code/michelman/Cognos Reports/Excel Validation/_validation_work/07 - Ivan SK 2023'
XLSX = r'C:/Users/Zack/Documents/Code/michelman/Cognos Reports/Excel Validation/1 - Ivan SK 2023.xlsx'

Q=OUT+'/'   # 2026-07-08: repointed to fresh full ADOMD pulls (_raw_pbi_*.csv), no 100-row cap
PBI = {
 'Inventory':          [Q+'_raw_pbi_Inventory.csv'],
 'Work_Orders':        [Q+'_raw_pbi_Work_Orders.csv'],
 'Sales_Order_Summary':[Q+'_raw_pbi_Sales_Order_Summary.csv'],
 'Inventory_HP':       [Q+'_raw_pbi_Inventory_HP.csv'],
 'Safety_Stock_HP':    [Q+'_raw_pbi_Safety_Stock_HP.csv'],
}
SHEET = {
 'Inventory':'Inventory_1','Work_Orders':'Work Order_2','Sales_Order_Summary':'Sales Orders_3',
 'Inventory_HP':'Inventory HP_4','Safety_Stock_HP':'Safety Stock HP_5',
}
# columns (by rendered position) that are numeric -> tolerance compare
NUMCOLS = {
 'Inventory':[9,10,14,15],           # QtyOnHand,HardCommit,OH KG,OH LB
 'Work_Orders':[10,11,14,15,16,17,18,19,20],
 'Sales_Order_Summary':[14,15,16,18],  # ORDER KGs, ORDER LBs, Prim QTY, 2nd QTY
 'Inventory_HP':[9,10],
 'Safety_Stock_HP':[6,7],
}
# date columns (by position) -> compare date-only
DATECOLS = {
 'Inventory':[13],
 'Work_Orders':[8,9,21],
 'Sales_Order_Summary':[20,21,22,23],
 'Inventory_HP':[],
 'Safety_Stock_HP':[],
}
# volatile "as of now" stamps (date/time of render) -> compare date-only, diff = cosmetic
VOLATILE = {'Inventory':[13],'Work_Orders':[21],'Sales_Order_Summary':[],'Inventory_HP':[],'Safety_Stock_HP':[]}
# diagnosis key columns (positions) per table
KEY = {
 'Inventory':[1,3,4,6,7,5,8],
 'Work_Orders':[3,6,12,8],
 'Sales_Order_Summary':[7,11,19,16],
 'Inventory_HP':[1,4,5,6,7],
 'Safety_Stock_HP':[1,4,5],
}

def norm_str(v):
    if v is None: return ''
    if isinstance(v,float) and v.is_integer(): v=int(v)
    return str(v).strip()

def norm_num(v):
    if v is None or (isinstance(v,str) and v.strip()==''): return None
    try: return float(str(v).replace(',',''))
    except: return None

def norm_date(v):
    if v is None: return ''
    if isinstance(v,(datetime.datetime,datetime.date)):
        return f'{v.year:04d}-{v.month:02d}-{v.day:02d}'
    s=str(v).strip()
    if s=='':return ''
    # parse "7/6/2026 12:00:00 AM"
    for fmt in ('%m/%d/%Y %I:%M:%S %p','%m/%d/%Y','%Y-%m-%d %H:%M:%S','%Y-%m-%d'):
        try:
            d=datetime.datetime.strptime(s,fmt); return f'{d.year:04d}-{d.month:02d}-{d.day:02d}'
        except: pass
    return s

def load_excel(sheet, ncol):
    wb=openpyxl.load_workbook(XLSX,data_only=True); ws=wb[sheet]
    rows=[]
    for r in range(2,ws.max_row+1):
        vals=[ws.cell(r,c).value for c in range(1,ncol+1)]
        if all(x is None or (isinstance(x,str) and x.strip()=='') for x in vals): continue
        rows.append(vals)
    return rows

def load_pbi(paths):
    hdr=None; rows=[]
    for path in paths:
        with open(path,encoding='utf-8-sig',newline='') as f:
            rd=csv.reader(f); h=next(rd)
            if hdr is None: hdr=h
            rows+=[r for r in rd]
    return hdr,rows

def canon(table, row, ncol):
    """return tuple of normalized values length ncol"""
    out=[]
    nums=set(NUMCOLS[table]); dates=set(DATECOLS[table])
    for i in range(ncol):
        v=row[i] if i<len(row) else None
        if i in nums:
            n=norm_num(v); out.append(None if n is None else round(n,2))
        elif i in dates:
            out.append(norm_date(v))
        else:
            out.append(norm_str(v))
    return out

def num_eq(a,b):
    if a is None and b is None: return True
    if a is None or b is None: return False
    return abs(a-b)<=0.02 or (max(abs(a),abs(b))>0 and abs(a-b)/max(abs(a),abs(b))<=1e-4)

def cell_eq(table,i,a,b):
    if i in set(NUMCOLS[table]): return num_eq(a,b)
    return a==b

summary=[]
resid_rows=[]
for table in ['Inventory','Work_Orders','Sales_Order_Summary','Inventory_HP','Safety_Stock_HP']:
    sheet=SHEET[table]; hdr,prows=load_pbi(PBI[table]); ncol=len(hdr)
    exrows=load_excel(sheet, ncol)  # excel may have extra trailing col (SSHP) -> handled: ncol=pbi cols
    # for SSHP excel has 9 cols; we read first ncol(8) only
    C=[canon(table,r,ncol) for r in exrows]
    P=[canon(table,r,ncol) for r in prows]
    # write per-side csvs
    with open(os.path.join(OUT,f'cognos_{table}.csv'),'w',newline='',encoding='utf-8') as f:
        w=csv.writer(f); w.writerow(hdr); w.writerows(C)
    with open(os.path.join(OUT,f'pbi_{table}.csv'),'w',newline='',encoding='utf-8') as f:
        w=csv.writer(f); w.writerow(hdr); w.writerows(P)
    # multiset match (numeric rounded already to 2dp; exact tuple match)
    cC=Counter(tuple(r) for r in C); cP=Counter(tuple(r) for r in P)
    inter=cC & cP; matched=sum(inter.values())
    only_c=cC - cP; only_p=cP - cC
    # key-based diagnosis on residuals
    keyidx=KEY[table]
    def keyof(r): return tuple(r[i] for i in keyidx)
    resC=defaultdict(list); resP=defaultdict(list)
    # expand residual multisets back to rows
    rc=[];
    for row,cnt in only_c.items(): rc+= [list(row)]*cnt
    rp=[]
    for row,cnt in only_p.items(): rp+= [list(row)]*cnt
    for r in rc: resC[keyof(r)].append(r)
    for r in rp: resP[keyof(r)].append(r)
    colmiss=Counter()
    paired=0; onlyC_key=0; onlyP_key=0
    allkeys=set(resC)|set(resP)
    for k in allkeys:
        lc=resC.get(k,[]); lp=resP.get(k,[])
        m=min(len(lc),len(lp))
        for j in range(m):
            paired+=1
            for i in range(ncol):
                if not cell_eq(table,i,lc[j][i],lp[j][i]):
                    colmiss[hdr[i]]+=1
                    resid_rows.append([table,'DIFF',hdr[i],str(k),repr(lc[j][i]),repr(lp[j][i])])
        for j in range(m,len(lc)):
            onlyC_key+=1; resid_rows.append([table,'COGNOS_ONLY','', str(k), '|'.join(map(str,lc[j])),''])
        for j in range(m,len(lp)):
            onlyP_key+=1; resid_rows.append([table,'PBI_ONLY','', str(k),'', '|'.join(map(str,lp[j]))])
    # build comparison csv (per residual) already in resid_rows; also write matched summary
    # comparison_<table>.csv : row-level status
    cPcopy=Counter(cP)
    comp=[]
    for r in C:
        t=tuple(r)
        if cPcopy[t]>0:
            cPcopy[t]-=1; comp.append(['MATCH']+list(r))
        else:
            comp.append(['COGNOS_ONLY']+list(r))
    for t,c in cPcopy.items():
        for _ in range(c): comp.append(['PBI_ONLY']+list(t))
    with open(os.path.join(OUT,f'comparison_{table}.csv'),'w',newline='',encoding='utf-8') as f:
        w=csv.writer(f); w.writerow(['status']+hdr); w.writerows(comp)
    rate = matched/max(len(C),1)*100
    summary.append((table,len(C),len(P),matched,round(rate,2),paired,onlyC_key,onlyP_key,dict(colmiss)))
    print(f'--- {table}: Cognos={len(C)} PBI={len(P)} exact_match={matched} ({rate:.2f}%)')
    print(f'    residual: paired_with_celldiffs={paired} cognos_only={onlyC_key} pbi_only={onlyP_key}')
    if colmiss: print('    col mismatches:',dict(colmiss))

with open(os.path.join(OUT,'residuals.csv'),'w',newline='',encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['table','type','column','key','cognos','pbi']); w.writerows(resid_rows)

print('\n==== SUMMARY ====')
for s in summary:
    print(s[0],'Cognos',s[1],'PBI',s[2],'match',s[3],f'{s[4]}%','| pairedDiffs',s[5],'Conly',s[6],'Ponly',s[7])
