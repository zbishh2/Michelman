import openpyxl, csv, os, glob

OUT = "_validation_work/01 - RM Staging"
os.makedirs(OUT, exist_ok=True)

# ---------- 1. Parse Cognos export (page 1 top table) ----------
wb = openpyxl.load_workbook('DEMO - RM Staging at Shell Road 2026.xlsx', data_only=True)
ws = wb['page']
cog_headers = [ws.cell(1,c).value for c in range(1,5)]
cog_rows = []
for r in range(2, ws.max_row+1):
    rm = ws.cell(r,1).value
    if rm is None: continue
    cog_rows.append({
        'RM': str(rm).strip(),
        'QTY OH in CINC': ws.cell(r,2).value,
        'Total RM Needed': ws.cell(r,3).value,
        'QTY Required from CIN2': ws.cell(r,4).value,
    })

with open(f"{OUT}/cognos_page.csv","w",newline="") as f:
    w = csv.writer(f); w.writerow(['RM','QTY OH in CINC','Total RM Needed','QTY Required from CIN2'])
    for row in cog_rows:
        w.writerow([row['RM'],row['QTY OH in CINC'],row['Total RM Needed'],row['QTY Required from CIN2']])

# ---------- 2. Load PBI table dumps from MCP CSV results ----------
def load_mcp_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rdr = csv.reader(f)
        rows = list(rdr)
    hdr = [h.split('[')[-1].rstrip(']') for h in rows[0]]
    return hdr, rows[1:]

qdir = os.path.expandvars(r"C:/Users/Zack/AppData/Local/Temp/PowerBIModelingMCP/QueryResults")
files = sorted(glob.glob(qdir+"/dax_query_result_*.csv"), key=os.path.getmtime)
# identify RM Requirements dump and WorkOrder dump by header
rm_pbi=None; wo_pbi=None
for fp in files:
    hdr, data = load_mcp_csv(fp)
    if hdr and hdr[0]=='RM' and 'Qty On Hand CINC' in hdr:
        rm_pbi=(hdr,data)
    if hdr and 'Work Order Start' in hdr:
        wo_pbi=(hdr,data)

# write pbi dumps
with open(f"{OUT}/pbi_RM Requirements.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(rm_pbi[0])
    for row in rm_pbi[1]: w.writerow(row)
with open(f"{OUT}/pbi_WorkOrder_Detail.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(wo_pbi[0])
    for row in wo_pbi[1]: w.writerow(row)

# ---------- 3. Compare RM Requirements top table ----------
def num(v):
    if v is None or v=='' : return None
    try: return float(v)
    except: return None

pbi_map={}
for row in rm_pbi[1]:
    d=dict(zip(rm_pbi[0],row))
    pbi_map[d['RM'].strip()]=d

cols=[('QTY OH in CINC','Qty On Hand CINC'),
      ('Total RM Needed','Total RM Needed'),
      ('QTY Required from CIN2','Qty Required From CIN2')]

comp_rows=[]
all_keys=sorted(set([r['RM'] for r in cog_rows])|set(pbi_map.keys()))
TOL=0.01
fully=0
colmiss={c[0]:0 for c in cols}
residuals=[]
for k in all_keys:
    cog=next((r for r in cog_rows if r['RM']==k),None)
    pbi=pbi_map.get(k)
    if cog is None:
        residuals.append(['PBI-only',k,'in PBI RM Requirements but not in Cognos export']); continue
    if pbi is None:
        residuals.append(['Cognos-only',k,'in Cognos export but not in PBI RM Requirements']); continue
    rowout={'RM':k,'RM_match':1}
    rowmatch=True
    for ccol,pcol in cols:
        cv=num(cog[ccol]); pv=num(pbi[pcol])
        if cv is None and pv is None:
            flag=1
        elif cv is None or pv is None:
            flag=0
        else:
            flag=1 if abs(cv-pv)<=TOL else 0
        rowout[f'cog_{ccol}']=cog[ccol]
        rowout[f'pbi_{ccol}']=pbi[pcol]
        rowout[f'{ccol}_match']=flag
        if flag==0: colmiss[ccol]+=1; rowmatch=False
    if rowmatch: fully+=1
    comp_rows.append(rowout)

# write comparison
fieldnames=['RM','RM_match']
for ccol,_ in cols:
    fieldnames+=[f'cog_{ccol}',f'pbi_{ccol}',f'{ccol}_match']
with open(f"{OUT}/comparison_rm_requirements.csv","w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=fieldnames); w.writeheader()
    for r in comp_rows: w.writerow(r)

with open(f"{OUT}/residuals.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(['side','key','diagnosis'])
    for r in residuals: w.writerow(r)

print("Cognos rows:",len(cog_rows))
print("PBI RM Requirements rows:",len(rm_pbi[1]))
print("PBI WorkOrder_Detail rows:",len(wo_pbi[1]))
print("Fully matched rows:",fully,"/",len(comp_rows))
print("Per-column mismatches:",colmiss)
print("Residuals:",residuals)
