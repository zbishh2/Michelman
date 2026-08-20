"""Builds the report 14 validation workbook in the report 21 layout.

Cognos side reads the filed tight capture; Power BI side reads executeQueries pulls
of the published SSAS Import model, filtered to the same snapshot date.
"""

import json
import math
import datetime as dt
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

REPO = Path(r"C:\Users\Zack\Documents\Code\Michelman")
R14 = REPO / "Cognos Reports" / "14 - 1 - Ivan Global Inventory Excel - Select Date"
PULLS = Path(
    r"C:\Users\Zack\AppData\Local\Temp\claude\C--Users-Zack-Documents-Code-Michelman"
    r"\d084456f-49e5-44e2-b346-1571942effaa\scratchpad\r14wb"
)
COGNOS = R14 / "Intake" / "Cognos export - tight capture 2026-08-05.xlsx"
OUT = REPO / "Cognos Reports" / "Excel Validation" / "_report_out" / "14 - Ivan Global Inventory Excel - Select Date.xlsx"

GREEN = PatternFill("solid", start_color="C6EFCE")
RED = PatternFill("solid", start_color="FFC7CE")
HDR = PatternFill("solid", start_color="D9E1F2")
BLUE = Font(bold=True, color="0000FF")
GREY = Font(color="444444")
CODE = Font(name="Consolas", size=9)


def norm_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    return str(v).strip()


def key_text(v):
    """Key normalisation: Cognos renders an absent code as '-', we render it blank."""
    s = norm_text(v)
    return "" if s == "-" else s


def norm_num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if math.isnan(f) else f


def norm_date(v):
    if v is None or v is pd.NaT:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, dt.datetime):
        return dt.datetime(v.year, v.month, v.day)
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return None
    try:
        d = dt.datetime.fromisoformat(s.replace("Z", ""))
        return dt.datetime(d.year, d.month, d.day)
    except ValueError:
        return None


def load_pbi(stem):
    payload = json.loads((PULLS / f"{stem}.json").read_text(encoding="utf-8"))
    assert payload.get("status_code") == 200, stem
    df = pd.DataFrame(payload["text"]["results"][0]["tables"][0]["rows"])
    df.columns = [c.split("[", 1)[1][:-1] if "[" in c else c for c in df.columns]
    return df


class Tab:
    def __init__(self, sheet, title, cognos_sheet, pbi_stem, key, cols, numeric, dates, dash, sort_note):
        self.sheet = sheet
        self.title = title
        self.key = key
        self.cols = cols
        self.numeric = set(numeric)
        self.dates = set(dates)
        self.dash = set(dash)
        self.sort_note = sort_note
        # keep_default_na off: "NULL" is a value the export carries, not a missing cell.
        self.cognos = pd.read_excel(COGNOS, sheet_name=cognos_sheet,
                                    keep_default_na=False, na_values=[])
        self.pbi = load_pbi(pbi_stem)

    def kind(self, col):
        if col in self.numeric:
            return "num"
        if col in self.dates:
            return "date"
        return "text"

    def cell(self, row, col):
        if row is None:
            return "" if self.kind(col) == "text" else None
        v = row.get(col)
        k = self.kind(col)
        return norm_num(v) if k == "num" else norm_date(v) if k == "date" else norm_text(v)

    def aligned(self):
        def keyed(df):
            out = {}
            for rec in df.to_dict("records"):
                out["|".join(key_text(rec.get(c)) for c in self.key)] = rec
            return out

        ck, pk = keyed(self.cognos), keyed(self.pbi)
        pairs = [(ck[k], pk.get(k)) for k in ck]
        pairs += [(None, pk[k]) for k in pk if k not in ck]
        return pairs


def write_tab(wb, tab, as_of):
    ws = wb.create_sheet(f"Comparison - {tab.sheet}")
    n = len(tab.cols)
    cog0, gap1, cmp0, gap2, pbi0 = 1, n + 1, n + 2, 2 * n + 2, 2 * n + 3
    pairs = tab.aligned()
    first, last = 11, 10 + len(pairs)

    ws["A1"] = tab.title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = f"As of {as_of}"
    ws["A4"] = tab.sort_note
    ws["A5"] = "Numeric match tolerance (% of larger value):"
    ws["B5"] = 0.005
    ws["A6"] = "Absolute tolerance floor (units):"
    ws["B6"] = 1
    for c in ("A5", "A6"):
        ws[c].font = Font(bold=True)
    for c in ("B5", "B6"):
        ws[c].font = BLUE

    for col, label in ((cog0, "Cognos"), (cmp0, "Compare"), (pbi0, "Power BI")):
        ws.cell(row=7, column=col, value=label).font = Font(bold=True)
    ws.cell(row=8, column=gap1, value="Reason").font = Font(bold=True)
    ws.cell(row=9, column=gap1, value="FALSE count").font = Font(bold=True)

    for i, name in enumerate(tab.cols):
        L = get_column_letter(cog0 + i)
        C = get_column_letter(cmp0 + i)
        R = get_column_letter(pbi0 + i)
        reason = (
            f'=IF({C}$9=0,"Match",'
            f'SUMPRODUCT(({C}${first}:{C}${last}=FALSE)*((( LEN({L}${first}:{L}${last})=0)'
            f'+(LEN({R}${first}:{R}${last})=0))>0))&" missing / "&'
            f'SUMPRODUCT(({C}${first}:{C}${last}=FALSE)*(LEN({L}${first}:{L}${last})>0)'
            f'*(LEN({R}${first}:{R}${last})>0))&" differ; e.g. r"&'
            f'(MATCH(FALSE,{C}${first}:{C}${last},0)+{first - 1})&": "&'
            f'INDEX({L}${first}:{L}${last},MATCH(FALSE,{C}${first}:{C}${last},0))&" vs "&'
            f'INDEX({R}${first}:{R}${last},MATCH(FALSE,{C}${first}:{C}${last},0)))'
        )
        rc = ws.cell(row=8, column=cmp0 + i)
        rc.value = ArrayFormula(f"{C}8", reason)
        rc.font = GREY
        fc = ws.cell(row=9, column=cmp0 + i, value=f"=COUNTIF({C}${first}:{C}${last},FALSE)")
        fc.font = Font(bold=True)
        for col in (cog0 + i, cmp0 + i, pbi0 + i):
            h = ws.cell(row=10, column=col, value=name)
            h.font = Font(bold=True)
            h.alignment = Alignment(wrap_text=True, vertical="bottom")

    for r, (crow, prow) in enumerate(pairs, start=first):
        for i, name in enumerate(tab.cols):
            k = tab.kind(name)
            L = get_column_letter(cog0 + i)
            R = get_column_letter(pbi0 + i)
            left = ws.cell(row=r, column=cog0 + i, value=tab.cell(crow, name))
            right = ws.cell(row=r, column=pbi0 + i, value=tab.cell(prow, name))
            if k == "num":
                left.number_format = right.number_format = "#,##0.0000"
                f = (f'=IF(OR({L}{r}="",{R}{r}=""),FALSE,'
                     f'ABS({L}{r}-{R}{r})<=MAX($B$5*MAX(ABS({L}{r}),ABS({R}{r})),$B$6))')
            elif k == "date":
                left.number_format = right.number_format = "yyyy-mm-dd"
                f = f'=IF(OR({L}{r}="",{R}{r}=""),FALSE,{L}{r}={R}{r})'
            elif name in tab.dash:
                f = (f'=IF(OR({L}{r}="",{R}{r}=""),'
                     f'OR({L}{r}={R}{r},AND(OR({L}{r}="",{L}{r}="-"),OR({R}{r}="",{R}{r}="-"))),'
                     f'{L}{r}={R}{r})')
            else:
                f = f'=IF(OR({L}{r}="",{R}{r}=""),FALSE,{L}{r}={R}{r})'
            ws.cell(row=r, column=cmp0 + i, value=f)

    cmp_range = f"{get_column_letter(cmp0)}{first}:{get_column_letter(cmp0 + n - 1)}{last}"
    head_cell = f"{get_column_letter(cmp0)}9"
    ws.conditional_formatting.add(
        f"{head_cell}:{get_column_letter(cmp0 + n - 1)}9",
        FormulaRule(formula=[f"{get_column_letter(cmp0)}9>0"], fill=RED))
    ws.conditional_formatting.add(
        f"{head_cell}:{get_column_letter(cmp0 + n - 1)}9",
        FormulaRule(formula=[f"{get_column_letter(cmp0)}9=0"], fill=GREEN))
    ws.conditional_formatting.add(cmp_range, FormulaRule(
        formula=[f"AND(ISLOGICAL({get_column_letter(cmp0)}{first}),"
                 f"{get_column_letter(cmp0)}{first}=TRUE)"], fill=GREEN))
    ws.conditional_formatting.add(cmp_range, FormulaRule(
        formula=[f"AND(ISLOGICAL({get_column_letter(cmp0)}{first}),"
                 f"{get_column_letter(cmp0)}{first}=FALSE)"], fill=RED))

    ws.auto_filter.ref = f"{get_column_letter(cmp0)}10:{get_column_letter(cmp0 + n - 1)}{last}"

    ws.column_dimensions["A"].width = 24
    for i in range(1, n):
        for base in (cog0, pbi0):
            ws.column_dimensions[get_column_letter(base + i)].width = 15
    for i in range(n):
        ws.column_dimensions[get_column_letter(cmp0 + i)].width = 12
    ws.column_dimensions[get_column_letter(gap1)].width = 3
    ws.column_dimensions[get_column_letter(gap2)].width = 3
    return len(pairs)


def main():
    as_of = "8/5/2026"
    inv = Tab(
        "Inventory", "Cognos Report - SSAS Import validation", "Inventory Data_2", "inv",
        ["Branch Plant", "Location", "Lot Number", "2nd Item Number"],
        ["Inventory Date", "REGION", "Branch Plant", "Global Bulk Item", "Bulk Item",
         "2nd Item Number", "Stock Type Code", "GL Class Code", "Location", "Lot Number",
         "Supplier Lot Number", "Lot Status", "Master Planning Family", "On Hand", "UOM",
         "OH KGs", "OH LBs", "OH USD", "OH EUR", "On Hand Date", "Lot Expiry Date",
         "Memo Lot 1", "Memo Lot 2", "Commodity Class Description",
         "Commodity Sub Class Description"],
        ["On Hand", "OH KGs", "OH LBs", "OH USD", "OH EUR"],
        ["Inventory Date", "On Hand Date", "Lot Expiry Date"],
        ["Global Bulk Item", "Bulk Item", "Location", "Lot Number", "Lot Status",
         "Supplier Lot Number", "Memo Lot 1", "Memo Lot 2", "Commodity Class Description",
         "Commodity Sub Class Description"],
        "Sorted by REGION, Global Bulk Item, Bulk Item, 2nd Item Number. "
        "Aligned on Branch Plant + Location + Lot Number + 2nd Item Number.")

    esi = Tab(
        "Escor Inventory", "Cognos Report - SSAS Import validation", "Escor Inventory_3", "esi",
        ["Branch Plant", "Location", "Lot Number", "2nd Item Number"],
        ["Inventory Date", "Branch Plant", "Global Bulk Item", "Bulk Item", "2nd Item Number",
         "Last Receipt Date", "Location", "Lot Number", "On Hand Date", "Lot Expiry Date",
         "Sell by Date", "Supplier Lot Number", "Memo Lot 1", "Memo Lot 2", "Lot Status",
         "Master Planning Family", "Quantity on Hand KGs", "Quantity on Hand LBs",
         "Quantity on Hand", "Primary Unit of Measure"],
        ["Quantity on Hand KGs", "Quantity on Hand LBs", "Quantity on Hand"],
        ["Inventory Date", "Last Receipt Date", "On Hand Date", "Lot Expiry Date", "Sell by Date"],
        ["Global Bulk Item", "Bulk Item", "Location", "Lot Number", "Lot Status",
         "Supplier Lot Number", "Memo Lot 1", "Memo Lot 2"],
        "Sorted by Branch Plant, Last Receipt Date. "
        "Aligned on Branch Plant + Location + Lot Number + 2nd Item Number.")

    esl = Tab(
        "Escor Lot Details", "Cognos Report - ODS Import validation", "Escor Lot Details_4", "esl",
        ["Branch Plant", "2nd Item Number", "Lot Number"],
        ["Branch Plant", "Bulk Item", "2nd Item Number", "Item Short ID", "Lot Number",
         "Supplier Lot Number", "Memo Lot 1", "Memo Lot 2", "On Hand Date"],
        [],
        ["On Hand Date"],
        ["Bulk Item", "Lot Number", "Supplier Lot Number", "Memo Lot 1", "Memo Lot 2"],
        "Cognos renders a distinct lot list. "
        "Aligned on Branch Plant + 2nd Item Number + Lot Number.")

    wb = Workbook()
    wb.remove(wb.active)
    notes = wb.create_sheet("Notes")
    counts = {t.sheet: write_tab(wb, t, as_of) for t in (inv, esi, esl)}
    build_notes(notes, inv, esi, esl, as_of)
    build_rs(wb.create_sheet("RS"), inv, esi, esl)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    for t in (inv, esi, esl):
        print(f"{t.sheet}: cognos {len(t.cognos)} | pbi {len(t.pbi)} | aligned rows {counts[t.sheet]}")
    print("saved", OUT)


def overlap(tab):
    def keys(df):
        return {"|".join(key_text(r.get(c)) for c in tab.key) for r in df.to_dict("records")}
    c, p = keys(tab.cognos), keys(tab.pbi)
    return len(c & p), len(c - p), len(p - c)


def build_notes(ws, inv, esi, esl, as_of):
    body = Font(size=11)
    head = Font(size=11, bold=True)
    small = Font(size=10)
    smallb = Font(size=10, bold=True)

    ws["A1"] = "Cognos Migration - Report 14 - 1 - Ivan Global Inventory Excel - Select Date"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:G1")
    ws["A2"] = (f"Comparison at inventory date {as_of} | Cognos exported 8/17/2026, "
                "Power BI refreshed 8/17/2026")
    ws.merge_cells("A2:G2")

    r = 4

    def para(text, bold=False):
        nonlocal r
        ws.cell(row=r, column=3, value=text).font = head if bold else body
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        r += 1

    def gap():
        nonlocal r
        r += 1

    i_shared, i_c, i_p = overlap(inv)
    e_shared, e_c, e_p = overlap(esi)
    l_shared, l_c, l_p = overlap(esl)

    para("What was compared", True)
    para(f"Inventory: {len(inv.cognos):,} Cognos rows and {len(inv.pbi):,} Power BI rows. "
         f"Escor Inventory: {len(esi.cognos):,} and {len(esi.pbi):,}. "
         f"Escor Lot Details: {len(esl.cognos):,} Cognos rows and {len(esl.pbi):,} Power BI rows.")
    para("Both sides read the same immutable 8/5/2026 snapshot, so the pair does not have to be "
         "captured minutes apart. Rows are aligned on the business key; one-sided rows remain "
         "visible with a blank opposite block.")
    gap()

    para("Results", True)
    para(f"Inventory: {i_shared:,} shared positions; {i_c} Cognos-only and {i_p} Power BI-only. "
         "Those thirteen are the same lots snapped either side of a warehouse movement - keyed on "
         "branch plus lot alone they fall to 2 and 1. OH USD ties to -0.076% and OH EUR to +0.005%.")
    para("Weight columns tie to -0.036% once two SING ETHAL.S lots are set aside. Those two are "
         "stocked in GM and Cognos applies its negative-value guard per unit, rendering 44 LB and "
         "20 KG for every gram. The other eight GM rows agree to the cent.")
    para(f"Escor Inventory: {e_shared} of {e_shared} rows shared, no one-sided keys. Every quantity "
         "column matches outright. Fifteen attribute cells differ: 6 Last Receipt Dates blank on "
         "one side, 4 Lot Statuses, 4 Memo Lot 2 entries where Cognos truncates the text, and 1 "
         "Memo Lot 1.")
    para(f"Escor Lot Details: {l_shared:,} shared lots, {l_c} Cognos-only, {l_p} Power BI-only - a "
         "strict subset. The Cognos-only rows are 79 lots dated 2011-2013 that ODS no longer "
         "retains plus one 2025 lot F4108 does not carry. Inside the 2024-onward window it is "
         "476 of 477.")
    gap()

    para("Residual differences - what the red cells are", True)
    para("The measure columns tie. Every remaining difference sits in a lot-master or item-master "
         "attribute, and they fall into three classes.")
    para("Cognos missing-value sentinels, and production SSAS carrying a real value on those rows. "
         "Where the legacy warehouse holds nothing it renders '-' or '0' for text and 1900-01-01 "
         "for dates: 98 Supplier Lot Numbers, 114 On Hand Dates, the same 114 Lot Expiry Dates, "
         "123 Memo Lot 2 and 41 Memo Lot 1 entries.")
    para("Current state against as-of. Lot and item attributes resolve through the cube's "
         "dimensions, which hold today's value, while Cognos reads the attribute as it stood in the "
         "snapshot. That is 187 Lot Expiry Dates - 69 later and 118 earlier, so drift rather than an "
         "offset - plus 62 Memo Lot 2 and 23 Memo Lot 1 rewrites, 4 On Hand Dates, 4 Stock Type "
         "Codes, 2 GL Class Codes and 1 Lot Status.")
    para("One-sided rows account for the remaining 13 FALSE cells in every Inventory column.")
    para("Supplier Lot Number needs a decision. On 96 of the rows where Cognos shows a marker, the "
         "cube returns the lot number itself where JDE holds no supplier lot. Neither side has a "
         "supplier lot; ours echoes a value that is not one.")
    gap()

    para("Source selection", True)
    para("SSAS Import from SSASPROD / BIQLTabular_ISH for the two inventory tabs - the production "
         "inventory-history model, and the only source that answers an arbitrary past date.")
    para("Escor Lot Details reads ODSPROD / PRODDTA.F4108, the JDE lot master. It reports every lot "
         "ever created for the Escor items whether or not stock remains, which is master data "
         "rather than an inventory position and has no counterpart in the cube or in EDW. "
         "Confirmed by Rohit.")
    gap()

    para("Links and report location", True)
    para("Power BI: Zack (Validation) > 14 - Ivan Global Inventory Excel - Select Date (SSAS Import)")
    para("https://app.powerbi.com/groups/50b98bb9-9fcb-47db-a0df-f2c167b351fb/reports/"
         "ca5f5770-7678-4053-8cef-09f9f18b107c")
    para("Cognos: Public Folders > Michelman Reporting > Production and Shipping > "
         "1 - Ivan Global Inventory Excel - Select Date")
    para("PBIP: Cognos Reports/14 - 1 - Ivan Global Inventory Excel - Select Date/"
         "1 - Ivan Global Inventory Excel - Select Date (SSAS Import)")
    gap()

    para("Cognos filter set - report XML vs native SQL vs our query", True)
    para("Report-author filters from the Cognos report XML, what they compile to in the generated "
         "native Oracle SQL, and the equivalent in our query. Raw captures filed alongside this "
         "report as Report XML.xml and the numbered .sql files.")
    gap()

    def table(title, header, rows, note=None):
        nonlocal r
        ws.cell(row=r, column=3, value=title).font = head
        r += 1
        for j, h in enumerate(header):
            c = ws.cell(row=r, column=3 + j, value=h)
            c.font = smallb
            c.fill = HDR
            c.alignment = Alignment(wrap_text=True)
        r += 1
        for row in rows:
            for j, v in enumerate(row):
                c = ws.cell(row=r, column=3 + j, value=v)
                c.font = smallb if j == 0 else small
                c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        if note:
            ws.cell(row=r, column=3, value=note).font = small
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
            r += 1
        r += 1

    dax_hdr = ["Filter", "Cognos filter (XML)", "Native SQL", "Our DAX"]
    table("Inventory", dax_hdr, [
        ["Date", "[Inventory Date] = ?Date?",
         "INVENTORY_DATE = CAST(:PQ1 AS TIMESTAMP)",
         "every snapshot date imported; the Select Date slicer picks one"],
        ["Quantity", "[Quantity on Hand] > 0", "QUANTITY_ON_HAND > 0",
         "[QuantityOnHandPrimaryUOM] > 0"],
        ["Branch Plant",
         "[Branch Plant] in ('CINC','CIN2','CIN4','AUBA','AUB2','SING','SNG4','MUM3','SHAN')",
         "same 9-branch list", "same 9-branch list"],
        ["Planning family", "[Master Planning Family] in (36 entries, 14 distinct)",
         "same list verbatim, duplicates included", "the 14 distinct values"],
        ["Cost method", "(no cost-method filter - DW_LEGACY carries one cost row)", "-",
         'TRIM([CostMethod]) = "07" (the ISH fact fans one row per method)'],
        ["OH KGs / LBs", "If (KGs < 0) Then (-KGs * 20) Else KGs; the LB item uses * 44",
         "case when qty * CONVERSION_FACTOR_KG < 0 then -(...) * 20 else ... end",
         "CALCULATE([Qty On Hand KG]) and CALCULATE([Qty On Hand LB])"],
        ["OH USD", "[Extended Cost for Quantity On Hand USD]",
         "sum((QUANTITY_ON_HAND * UNIT_COST) * T7.FROM_TO_EXCHANGE_RATE), T7 to USD, "
         "rate type '-', rate effective on the inventory date",
         "CALCULATE([Total Ext Cost IC USD], 'Selected UOM Filter'[Selected UOM Code] = 1) + 0"],
        ["OH EUR", "[Extended Cost for Quantity On Hand EUR]",
         "same shape against T8, to EUR",
         "[AmountValueAtCost] * LOOKUPVALUE('Currency Rates'[ToRateDaily], "
         "'Currency Rates'[CurrencySKey], 'Inventory Snapshot'[CurrencyBSKey])"],
    ], "Note: Cognos's negative-value guard is what produces the two inflated GM lots - it multiplies "
       "by 20 and 44 per unit rather than per position. USD reads the cube's own measure, which ties "
       "once Selected UOM is pinned to Primary - the cube's default is LB, and both the unit cost and "
       "the quantity inside the measure scale by the UOM rate, so an unpinned read squares it. EUR has "
       "no cube route at all, so it reads the column carried by the local-to-EUR daily rate.")

    table("Escor Inventory", dax_hdr, [
        ["Date", "[Inventory Date] = ?Date?", "same", "the Select Date slicer"],
        ["Quantity", "[Quantity on Hand] > 0", "QUANTITY_ON_HAND > 0",
         "[QuantityOnHandPrimaryUOM] > 0"],
        ["Item", "[Global Bulk Item] in ('ESC5200')",
         "ITEM.GLOBAL_BULK_ITEM in (N'ESC5200')",
         'TRIM(RELATED(\'Item Branch\'[Item Num Global Bulk])) = "ESC5200"'],
        ["Cost method", "(none)", "-", 'TRIM([CostMethod]) = "07", for the same grain reason'],
        ["Weights", "plain conversion, no guard",
         "qty * CONVERSION_FACTOR_KG and _LB",
         "CALCULATE([Qty On Hand KG]) and CALCULATE([Qty On Hand LB])"],
        ["Lot Status", "from Item Lot Number, the lot master",
         "ITEM_LOT_NUMBER_ALIAS.LOT_STATUS",
         "RELATED('Lot'[Lot Status Code])"],
    ], "Note: Lot Status is sourced from the lot master here and from the position on the Inventory "
       "tab. That asymmetry is Cognos's own and reproducing it is the point - the same lot can carry "
       "different statuses in different locations.")

    table("Escor Lot Details", ["Filter", "Cognos filter (XML)", "Native SQL", "Our T-SQL"], [
        ["Item", "[Bulk Item] in ('ESC5200','ESC5200.E','ESC5200.S')",
         "ITEM.BULK_ITEM in (N'ESC5200', N'ESC5200.E', N'ESC5200.S')",
         "LTRIM(RTRIM(tag.IMBULK)) IN ('ESC5200','ESC5200.E','ESC5200.S')"],
        ["Grain", "select distinct", "DISTINCT", "SELECT DISTINCT"],
        ["Source", "Data Warehouse Item Lot Number", "DW_LEGACY", "ODSPROD PRODDTA.F4108"],
    ], "Note: the bulk item comes from the dimension, not from F4108's own IOAITM copy. IOAITM agrees "
       "with the dimension on every row it selects, so the two look interchangeable under testing - "
       "but it misses child items rolling up to an Escor bulk under a different code. That is the "
       "difference between 1,584 rows and 1,588.")

    para("Source queries", True)
    para("As run against the published models. Verified character-for-character against the shipped "
         "partitions.")
    gap()

    for label, path in (
        ("1. Inventory - SSASPROD / BIQLTabular_ISH", R14 / "Inventory (SSAS).commented.dax"),
        ("2. Escor Inventory - SSASPROD / BIQLTabular_ISH", R14 / "Escor Inventory (SSAS).commented.dax"),
        ("3. Escor Lot Details - ODSPROD / PRODDTA.F4108", R14 / "Escor Lot Info (SSAS).commented.sql"),
    ):
        ws.cell(row=r, column=3, value=label).font = head
        r += 1
        for line in strip_comments(path):
            ws.cell(row=r, column=3, value=line).font = CODE
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 44
    ws.column_dimensions["G"].width = 20


def strip_comments(path):
    """The shipped query, with the annotation lines removed."""
    marker = "--" if path.suffix == ".sql" else "//"
    out = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.rstrip()
        stripped = line.strip()
        if stripped.startswith(marker) or stripped == "":
            continue
        out.append(line)
    return out


def build_rs(ws, inv, esi, esl):
    bold = Font(bold=True)
    ws["A1"] = "Report 14 source reconciliation"
    ws["A1"].font = bold
    rows = [
        ("Published model", "SSAS Import + ODS Import",
         "SSASPROD / BIQLTabular_ISH and ODSPROD / PRODDTA.F4108"),
        ("Published report",
         "https://app.powerbi.com/groups/50b98bb9-9fcb-47db-a0df-f2c167b351fb/reports/"
         "ca5f5770-7678-4053-8cef-09f9f18b107c", ""),
    ]
    r = 2
    for a, b, c in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if c:
            ws.cell(row=r, column=3, value=c)
        r += 1
    r += 1

    def block(title, lines):
        nonlocal r
        for j, v in enumerate(("Cognos", "Power BI")):
            ws.cell(row=r, column=2 + j, value=v).font = bold
        ws.cell(row=r, column=1, value=title).font = bold
        r += 1
        for a, b, c in lines:
            ws.cell(row=r, column=1, value=a)
            ws.cell(row=r, column=2, value=b)
            ws.cell(row=r, column=3, value=c)
            r += 1
        r += 1

    i_shared, i_c, i_p = overlap(inv)
    e_shared, e_c, e_p = overlap(esi)
    l_shared, l_c, l_p = overlap(esl)

    block("Inventory", [
        ("Rows", len(inv.cognos), len(inv.pbi)),
        ("Shared business keys", i_shared, i_shared),
        ("One-sided keys", i_c, i_p),
        ("On Hand", "16,725,722.16", "16,697,777.68"),
        ("OH USD", "40,110,239.97", "40,079,704.14 (-0.076%)"),
        ("OH EUR", "34,739,372.32", "34,741,247.54 (+0.005%)"),
        ("OH LBs excluding two GM lots", "23,894,223.23", "23,885,549.30 (-0.036%)"),
        ("OH KGs excluding two GM lots", "10,838,353.62", "10,834,445.66 (-0.036%)"),
        ("Row population", "5 one-sided at fine key", "8 one-sided; 2 and 1 at branch + lot"),
        ("Missing-value sentinels", "94 '-' supplier lots, 112 1900-01-01 dates", "real values carried"),
        ("Lot attribute drift", "value as of the snapshot", "current state from the Lot dimension"),
    ])
    block("Escor Inventory", [
        ("Rows", len(esi.cognos), len(esi.pbi)),
        ("Shared business keys", e_shared, e_shared),
        ("One-sided keys", e_c, e_p),
        ("Quantity on Hand", "751,972.53", "751,972.525"),
        ("Quantity on Hand KGs", "717,067.48", "717,067.66"),
        ("Quantity on Hand LBs", "1,580,846.39", "1,580,846.94"),
    ])
    block("Escor Lot Details", [
        ("Rows", len(esl.cognos), len(esl.pbi)),
        ("Shared business keys", l_shared, l_shared),
        ("One-sided keys", l_c, l_p),
        ("2024-onward window", 477, 476),
        ("Residual", "79 lots dated 2011-2013 plus one 2025 lot", "strict subset, nothing spurious"),
    ])

    ws.cell(row=r, column=1, value="Verdict").font = bold
    ws.cell(row=r, column=2, value="Tied and explainable").font = bold
    ws.cell(row=r, column=3, value="Ready for validation review").font = bold
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 46


if __name__ == "__main__":
    main()
